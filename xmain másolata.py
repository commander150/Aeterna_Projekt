import random
import time
import os
import traceback
from datetime import datetime
from openpyxl import load_workbook

# =========================
# ELÉRÉSI ÚT MEGHATÁROZÁSA (Golyóálló verzió)
# =========================
PROGRAM_MAPPA = os.path.dirname(os.path.abspath(__file__))
XLSX_NEV = "cards.xlsx"
XLSX_TELJES_UT = os.path.join(PROGRAM_MAPPA, XLSX_NEV)
LOG_FAJL = os.path.join(PROGRAM_MAPPA, "aeterna_teszt_naplo.txt")

def naplozz(uzenet):
    idobelyeg = datetime.now().strftime("%H:%M:%S")
    sor = f"[{idobelyeg}] {uzenet}"
    print(sor)
    try:
        with open(LOG_FAJL, "a", encoding="utf-8") as f:
            f.write(sor + "\n")
    except:
        pass

# =========================
# ADATMODELL
# =========================
class Kartya:
    def __init__(self, sor_adat):
        self.nev = str(sor_adat[0]) if sor_adat[0] else "Névtelen"
        self.kartyatipus = str(sor_adat[1])
        self.birodalom = str(sor_adat[2])
        self.klan = str(sor_adat[3])
        self.faj = str(sor_adat[4])
        self.kaszt = str(sor_adat[5])
        
        def tisztit_szam(v):
            try: return int(float(v)) if v is not None else 0
            except: return 0

        self.magnitudo = tisztit_szam(sor_adat[6])
        self.aura_koltseg = tisztit_szam(sor_adat[7])
        self.tamadas = tisztit_szam(sor_adat[8])
        self.eletero = tisztit_szam(sor_adat[9])
        self.kepesseg = str(sor_adat[10]) if len(sor_adat) > 10 else ""

        self.egyseg_e = "Entitás" in self.kartyatipus
        self.jel_e = "Jel" in self.kartyatipus
        self.reakcio_e = "Reakció" in self.kepesseg or "Burst" in self.kepesseg

class CsataEgyseg:
    def __init__(self, kartya):
        self.lap = kartya
        self.akt_tamadas = kartya.tamadas
        self.akt_hp = kartya.eletero
        self.kimerult = True
        if "Gyorsaság" in kartya.kepesseg:
            self.kimerult = False

    def serul(self, mennyiseg):
        self.akt_hp -= mennyiseg
        return self.akt_hp <= 0

class Jatekos:
    def __init__(self, nev, birodalom_neve, teljes_kartyatar):
        self.nev = nev
        self.birodalom = birodalom_neve
        sajat = [k for k in teljes_kartyatar if k.birodalom == birodalom_neve]
        if not sajat:
            sajat = [k for k in teljes_kartyatar if k.birodalom in [birodalom_neve, "Aether"]]
        if len(sajat) < 5:
            raise ValueError(f"Hiba: {birodalom_neve} birodalomhoz nincs elég kártya!")
        self.pakli = [random.choice(sajat) for _ in range(30)]
        random.shuffle(self.pakli)
        self.kez, self.pecsetek, self.osforras = [], [], []
        self.horizont, self.zenit = [None]*6, [None]*6

    def huzas(self):
        if self.pakli: self.kez.append(self.pakli.pop())

    def osforras_bovites(self):
        if self.kez: self.osforras.append({"lap": self.kez.pop(random.randrange(len(self.kez))), "hasznalt": False})

    def aura_visszatoltes(self):
        for o in self.osforras: o["hasznalt"] = False
        for i in range(6):
            if self.horizont[i]: self.horizont[i].kimerult = False
            if isinstance(self.zenit[i], CsataEgyseg): self.zenit[i].kimerult = False

    def elerheto_aura(self): return sum(1 for o in self.osforras if not o["hasznalt"])

    def fizet(self, koltseg):
        szabad = [o for o in self.osforras if not o["hasznalt"]]
        if len(szabad) >= koltseg:
            for i in range(koltseg): szabad[i]["hasznalt"] = True
            return True
        return False

# =========================
# SZIMULÁCIÓ MOTOR
# =========================
class AeternaSzimulacio:
    def __init__(self, kartyak):
        birodalmak = list(set(k.birodalom for k in kartyak if k.birodalom and k.birodalom != "None"))
        b1 = random.choice(birodalmak)
        b2 = random.choice(birodalmak)
        self.p1 = Jatekos("Játékos_1", b1, kartyak)
        self.p2 = Jatekos("Játékos_2", b2, kartyak)
        self.kor = 1
        self.elokeszites()

    def elokeszites(self):
        for _ in range(5): self.p1.huzas(); self.p2.huzas()
        for _ in range(5):
            if self.p1.pakli: self.p1.pecsetek.append(self.p1.pakli.pop())
            if self.p2.pakli: self.p2.pecsetek.append(self.p2.pakli.pop())

    def kijatszas(self, jatekos):
        lehetosegek = [l for l in jatekos.kez if l.aura_koltseg <= jatekos.elerheto_aura() and l.magnitudo <= len(jatekos.osforras)]
        if not lehetosegek: return
        lap = random.choice(lehetosegek)
        if lap.egyseg_e:
            zona = jatekos.horizont if random.random() < 0.7 else jatekos.zenit
            for i in range(6):
                if zona[i] is None:
                    if jatekos.fizet(lap.aura_koltseg):
                        jatekos.kez.remove(lap)
                        zona[i] = CsataEgyseg(lap)
                        naplozz(f"{jatekos.nev} kijátszotta: {lap.nev}")
                        break

    def harc(self, tamado, vedo):
        for i in range(6):
            egyseg = tamado.horizont[i]
            if egyseg and not egyseg.kimerult:
                egyseg.kimerult = True
                blokkolok = [e for e in vedo.horizont if e and not e.kimerult]
                if blokkolok:
                    b = random.choice(blokkolok)
                    if b.serul(egyseg.akt_tamadas): vedo.horizont[vedo.horizont.index(b)] = None
                    if egyseg.serul(b.akt_tamadas): tamado.horizont[i] = None
                else:
                    if vedo.pecsetek:
                        p = vedo.pecsetek.pop()
                        vedo.kez.append(p)
                        naplozz(f"PECSÉT FELTÖRT: {p.nev}")
                    else: return tamado.nev
        return None

    def futtatas(self):
        while self.kor < 100:
            for akt, ell in [(self.p1, self.p2), (self.p2, self.p1)]:
                akt.aura_visszatoltes(); akt.huzas(); akt.osforras_bovites()
                self.kijatszas(akt)
                eredmeny = self.harc(akt, ell)
                if eredmeny: return eredmeny
            self.kor += 1
        return "Időtúllépés"

# =========================
# INDÍTÁS ÉS HIBAKERESÉS
# =========================
if __name__ == "__main__":
    try:
        print(f"DEBUG: Program mappája: {PROGRAM_MAPPA}")
        print(f"DEBUG: Itt keresem a fájlt: {XLSX_TELJES_UT}")

        if not os.path.exists(XLSX_TELJES_UT):
            print("\n!!! HIBA: A FÁJL NEM TALÁLHATÓ !!!")
            print(f"Győződj meg róla, hogy a '{XLSX_NEV}' fájl pontosan ebben a mappában van:")
            print(f"--> {PROGRAM_MAPPA}")
        else:
            naplozz("Kezdés: Excel betöltése...")
            wb = load_workbook(XLSX_TELJES_UT, data_only=True)
            osszes_kartya = []
            for lap_nev in wb.sheetnames:
                sheet = wb[lap_nev]
                for sor in sheet.iter_rows(min_row=2, values_only=True):
                    if sor and sor[0]: osszes_kartya.append(Kartya(sor))
            
            if not osszes_kartya:
                print("HIBA: Az Excel üres vagy nem sikerült kártyákat beolvasni!")
            else:
                szim = AeternaSzimulacio(osszes_kartya)
                print(f"Meccs eredménye: {szim.futtatas()}")
                print(f"Napló: {LOG_FAJL}")

    except Exception:
        print("\n" + "!"*40)
        print("PROGRAM LEÁLLT - HIBAÜZENET:")
        traceback.print_exc()
        print("!"*40)

    print("\n")
    input("Nyomj Entert a bezáráshoz...")