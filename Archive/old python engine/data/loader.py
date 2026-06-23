import json
import os
import re
from openpyxl import load_workbook
from engine.card_metadata import (
    DURATION_ALIASES,
    EFFECT_TAG_ALIASES,
    LEGACY_INTERNAL_VALUES,
    STANDARD_FIELD_VALUES,
    TARGET_ALIASES,
    TRIGGER_ALIASES,
    ZONE_ALIASES,
    normalize_condition_value,
    normalize_duration_value,
    normalize_effect_tag_value,
    normalize_status_value,
    normalize_target_value,
    normalize_trigger_value,
    normalize_zone_value,
    normalized_metadata_list,
)
from utils.text import normalize_lookup_text
from utils.logger import naplo
from engine.card import Kartya


HEADER_ALIASES = {
    "card_id": "card_id",
    "generated_print_id": "generated_print_id",
    "print_id": "print_id",
    "set_id": "set_id",
    "collector_number": "collector_number",
    "rarity": "rarity",
    "treatment": "treatment",
    "art_variant": "art_variant",
    "print_status": "print_status",
    "version": "version",
    "reprint_of": "reprint_of",
    "play_legal_status": "play_legal_status",
    "szabalyi_kartya_id": "card_id",
    "kartya_nev": "kartya_nev",
    "tipus": "kartyatipus",
    "kartyatipus": "kartyatipus",
    "birodalom": "birodalom",
    "klan": "klan",
    "faj": "faj",
    "kaszt": "kaszt",
    "magnitudo": "magnitudo",
    "aura": "aura_koltseg",
    "aura_koltseg": "aura_koltseg",
    "atk": "tamadas",
    "tamadas": "tamadas",
    "hp": "eletero",
    "eletero": "eletero",
    "kepesseg": "kepesseg",
    "kepesseg_canonical": "kepesseg_canonical",
    "zona_felismerve": "zona_felismerve",
    "kulcsszavak_felismerve": "kulcsszavak_felismerve",
    "trigger_felismerve": "trigger_felismerve",
    "celpont_felismerve": "celpont_felismerve",
    "hatascimkek": "hatascimkek",
    "idotartam_felismerve": "idotartam_felismerve",
    "feltetel_felismerve": "feltetel_felismerve",
    "gepi_leiras": "gepi_leiras",
    "ertelmezesi_statusz": "ertelmezesi_statusz",
    "engine_megjegyzes": "engine_megjegyzes",
}

STANDARD_COLUMN_ORDER = [
    "kartya_nev",
    "kartyatipus",
    "birodalom",
    "klan",
    "faj",
    "kaszt",
    "magnitudo",
    "aura_koltseg",
    "tamadas",
    "eletero",
    "kepesseg",
    "kepesseg_canonical",
    "zona_felismerve",
    "kulcsszavak_felismerve",
    "trigger_felismerve",
    "celpont_felismerve",
    "hatascimkek",
    "idotartam_felismerve",
    "feltetel_felismerve",
    "gepi_leiras",
    "ertelmezesi_statusz",
    "engine_megjegyzes",
]

LIST_FIELDS = {
    "zona_felismerve",
    "kulcsszavak_felismerve",
    "trigger_felismerve",
    "celpont_felismerve",
    "hatascimkek",
}

INTEGER_FIELDS = {
    "magnitudo",
    "aura_koltseg",
    "tamadas",
    "eletero",
}

RUNTIME_PRINT_COMPONENTS = (
    "set_id",
    "card_id",
    "rarity",
    "treatment",
    "art_variant",
    "print_status",
    "version",
)

RUNTIME_REQUIRED_FIELDS = {
    "card_id",
    "set_id",
    "rarity",
    "treatment",
    "art_variant",
    "print_status",
    "version",
}

RUNTIME_ENUM_VALUES = {
    "rarity": {"C", "UC", "R", "SR", "MR", "UR", "P"},
}

REQUIRED_TEXT_FIELDS = {
    "kartya_nev",
    "kartyatipus",
    "birodalom",
    "kepesseg",
}

ALLOWED_ZONES = set(STANDARD_FIELD_VALUES["zone"])
ALLOWED_KEYWORDS = set(STANDARD_FIELD_VALUES["keyword"])
ALLOWED_TRIGGERS = set(STANDARD_FIELD_VALUES["trigger"])
ALLOWED_TARGETS = set(STANDARD_FIELD_VALUES["target"])
ALLOWED_EFFECT_TAGS = set(STANDARD_FIELD_VALUES["effect_tag"])
ALLOWED_DURATIONS = set(STANDARD_FIELD_VALUES["duration"])

FIELD_NORMALIZERS = {
    "zona_felismerve": normalize_zone_value,
    "trigger_felismerve": normalize_trigger_value,
    "celpont_felismerve": normalize_target_value,
    "hatascimkek": normalize_effect_tag_value,
    "idotartam_felismerve": normalize_duration_value,
}

FIELD_ALIASES = {
    "zona_felismerve": ZONE_ALIASES,
    "trigger_felismerve": TRIGGER_ALIASES,
    "celpont_felismerve": TARGET_ALIASES,
    "hatascimkek": EFFECT_TAG_ALIASES,
    "idotartam_felismerve": DURATION_ALIASES,
}

FIELD_STANDARD_VALUES = {
    "zona_felismerve": ALLOWED_ZONES,
    "kulcsszavak_felismerve": ALLOWED_KEYWORDS,
    "trigger_felismerve": ALLOWED_TRIGGERS,
    "celpont_felismerve": ALLOWED_TARGETS,
    "hatascimkek": ALLOWED_EFFECT_TAGS,
    "idotartam_felismerve": ALLOWED_DURATIONS,
}

FIELD_LEGACY_VALUES = {
    "trigger_felismerve": LEGACY_INTERNAL_VALUES["trigger"],
    "hatascimkek": LEGACY_INTERNAL_VALUES["effect_tag"],
    "idotartam_felismerve": LEGACY_INTERNAL_VALUES["duration"],
}


def _normalize_header_name(header):
    normalized = normalize_lookup_text(header)
    normalized = normalized.replace(" ", "_")
    return HEADER_ALIASES.get(normalized, normalized)


def _normalize_jsonl_mapping(raw_mapping):
    return {_normalize_header_name(key): value for key, value in raw_mapping.items()}


def _row_to_mapping(headers, row):
    mapping = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        mapping[header] = row[index] if index < len(row) else None
    return mapping


def _normalize_scalar(field_name, value):
    if value is None:
        return "" if field_name not in INTEGER_FIELDS else 0

    if field_name in INTEGER_FIELDS:
        try:
            if str(value).strip() == "":
                return 0
            return int(float(value))
        except Exception:
            return 0

    text = str(value).strip()
    normalized = normalize_lookup_text(text)
    if normalized in {"blank", "none", "-", ""}:
        return ""
    if field_name == "idotartam_felismerve":
        durations = normalized_metadata_list(text, field_name="duration")
        return ", ".join(durations)
    if field_name == "feltetel_felismerve":
        return normalize_condition_value(text)
    if field_name == "ertelmezesi_statusz":
        return normalize_status_value(text)
    return text


def _normalize_list_value(field_name, value):
    if value is None:
        return ""
    if field_name == "zona_felismerve":
        items = normalized_metadata_list(value, field_name="zone")
    elif field_name == "kulcsszavak_felismerve":
        items = normalized_metadata_list(value)
    elif field_name == "trigger_felismerve":
        items = normalized_metadata_list(value, field_name="trigger")
    elif field_name == "celpont_felismerve":
        items = normalized_metadata_list(value, field_name="target")
    elif field_name == "hatascimkek":
        items = normalized_metadata_list(value, field_name="effect_tag")
    else:
        items = normalized_metadata_list(value)
    return ", ".join(items)


def _merge_list_token(csv_value, token):
    items = _split_normalized_csv(csv_value)
    normalized_token = normalize_lookup_text(token)
    if normalized_token and normalized_token not in items:
        items.append(normalized_token)
    return ", ".join(items)


def _split_normalized_csv(value):
    if not value:
        return []
    return [item.strip() for item in str(value).split(",") if item.strip()]


def classify_enum_token(field_name, token):
    normalized = normalize_lookup_text(token)
    allowed = FIELD_STANDARD_VALUES.get(field_name, set())
    if normalized in allowed:
        return ("standard", normalized)

    alias_map = FIELD_ALIASES.get(field_name, {})
    canonical = alias_map.get(normalized)
    if canonical in allowed:
        return ("alias_normalizable", canonical)

    if normalized in FIELD_LEGACY_VALUES.get(field_name, set()):
        canonical = FIELD_NORMALIZERS.get(field_name, lambda value: normalized)(normalized)
        return ("legacy_internal", canonical if canonical in allowed else "")

    if " " in str(token).strip():
        canonical = FIELD_NORMALIZERS.get(field_name, lambda value: normalized)(normalized)
        if canonical in allowed:
            return ("alias_normalizable", canonical)
        return ("invalid_delimiter_or_format", "")

    return ("unknown_enum_value", "")


def normalize_row_mapping(mapping):
    normalized = {}
    for field_name in STANDARD_COLUMN_ORDER:
        value = mapping.get(field_name)
        if field_name in LIST_FIELDS:
            normalized[field_name] = _normalize_list_value(field_name, value)
        else:
            normalized[field_name] = _normalize_scalar(field_name, value)

    if (
        normalize_lookup_text(normalized.get("kartyatipus", "")) == "jel"
        and normalize_lookup_text(normalized.get("idotartam_felismerve", "")) == "trap"
        and bool(normalized.get("trigger_felismerve"))
        and bool(normalized.get("hatascimkek"))
    ):
        # Legacy spreadsheet rows sometimes stored the trap card kind in the duration field.
        # For triggered Jel cards this is not a real duration, so we clear it here instead
        # of preserving a misleading unknown enum token.
        normalized["idotartam_felismerve"] = ""

    if normalize_lookup_text(normalized.get("kartyatipus", "")) == "ige":
        zones = _split_normalized_csv(normalized.get("zona_felismerve", ""))
        if "burst" in zones:
            zones = [zone for zone in zones if zone != "burst"]
            normalized["zona_felismerve"] = ", ".join(zones)
            normalized["kulcsszavak_felismerve"] = _merge_list_token(
                normalized.get("kulcsszavak_felismerve", ""),
                "burst",
            )
    return normalized


def _is_keyword_only_static_entity(mapping):
    return (
        normalize_lookup_text(mapping.get("kartyatipus", "")) == "entitas"
        and normalize_lookup_text(mapping.get("trigger_felismerve", "")) == "static"
        and normalize_lookup_text(mapping.get("idotartam_felismerve", "")) == "static_while_on_board"
        and not mapping.get("hatascimkek")
        and bool(mapping.get("kulcsszavak_felismerve"))
    )


def validate_row_mapping(mapping, row_index=None, sheet_name=None):
    issues = []
    for field_name in STANDARD_COLUMN_ORDER:
        if field_name not in mapping:
            issues.append(f"hianyzo_oszlop:{field_name}")
    for field_name in REQUIRED_TEXT_FIELDS:
        if not str(mapping.get(field_name, "")).strip():
            issues.append(f"ures_kotelezo_mezo:{field_name}")
    if mapping.get("kartyatipus") and normalize_lookup_text(mapping["kartyatipus"]) == "entitas":
        if int(mapping.get("eletero", 0) or 0) <= 0:
            issues.append("entitas_hp_hianyzik")
    enum_fields = (
        "zona_felismerve",
        "kulcsszavak_felismerve",
        "trigger_felismerve",
        "celpont_felismerve",
        "hatascimkek",
        "idotartam_felismerve",
    )
    for field_name in enum_fields:
        for token in _split_normalized_csv(mapping.get(field_name, "")):
            issue_type, canonical = classify_enum_token(field_name, token)
            if issue_type == "standard":
                continue
            if issue_type == "alias_normalizable":
                issues.append(f"alias_normalizable:{field_name}:{token}->{canonical}")
            elif issue_type == "legacy_internal":
                if canonical:
                    issues.append(f"legacy_internal_value:{field_name}:{token}->{canonical}")
                else:
                    issues.append(f"legacy_internal_value:{field_name}:{token}")
            elif issue_type == "invalid_delimiter_or_format":
                issues.append(f"invalid_delimiter_or_format:{field_name}:{token}")
            else:
                issues.append(f"unknown_enum_value:{field_name}:{token}")
    if normalize_lookup_text(mapping.get("kartyatipus", "")) == "entitas":
        suspicious_targets = {"enemy_spell", "enemy_spell_or_ritual", "own_hand", "enemy_hand"}
        for token in _split_normalized_csv(mapping.get("celpont_felismerve", "")):
            if normalize_lookup_text(token) in suspicious_targets:
                issues.append(f"suspicious_field_combination:gyanus_target_tipus_kombinacio:{token}")
    if (
        mapping.get("idotartam_felismerve")
        and not mapping.get("hatascimkek")
        and not _is_keyword_only_static_entity(mapping)
    ):
        issues.append("suspicious_field_combination:idotartam_hatascimke_nelkul")
    prefix = []
    if sheet_name:
        prefix.append(f"sheet={sheet_name}")
    if row_index is not None:
        prefix.append(f"row={row_index}")
    location = " ".join(prefix)
    return [f"{location} {issue}".strip() for issue in issues]


def generate_print_id(mapping):
    parts = []
    for field_name in RUNTIME_PRINT_COMPONENTS:
        value = mapping.get(field_name)
        text = str(value).strip() if value is not None else ""
        if not text or normalize_lookup_text(text) in {"blank", "none", "-"}:
            return ""
        parts.append(text)
    return "-".join(parts)


def _runtime_field_warnings(normalized, row_index):
    warnings = []
    card_id = str(normalized.get("card_id") or "").strip()
    if card_id and not re.fullmatch(r"[A-Z0-9]+-[A-Z0-9]+-[0-9]{3}", card_id):
        warnings.append(f"line={row_index} suspicious_field_format:card_id:{card_id}")

    art_variant = str(normalized.get("art_variant") or "").strip()
    if art_variant and not re.fullmatch(r"A[0-9]+", art_variant):
        warnings.append(f"line={row_index} suspicious_field_format:art_variant:{art_variant}")

    for field_name, allowed_values in RUNTIME_ENUM_VALUES.items():
        value = str(normalized.get(field_name) or "").strip()
        if value and value not in allowed_values:
            warnings.append(f"line={row_index} unknown_runtime_enum_value:{field_name}:{value}")
    return warnings


def _normalize_runtime_row(raw_mapping, row_index):
    warnings = []
    if not isinstance(raw_mapping, dict):
        return None, [f"line={row_index} invalid_jsonl_row:not_object"]

    first_key = next(iter(raw_mapping), None)
    if first_key != "Card_ID":
        warnings.append(f"line={row_index} suspicious_field_order:first_field_not_Card_ID:{first_key}")

    runtime_mapping = _normalize_jsonl_mapping(raw_mapping)
    if "print_id" in runtime_mapping:
        warnings.append(f"line={row_index} ignored_stored_print_id")
        runtime_mapping.pop("print_id", None)

    normalized = normalize_row_mapping(runtime_mapping)
    for field_name, value in runtime_mapping.items():
        if field_name not in normalized and field_name != "print_id":
            normalized[field_name] = value

    for field_name in RUNTIME_REQUIRED_FIELDS:
        value = normalized.get(field_name)
        if value is None or not str(value).strip():
            warnings.append(f"line={row_index} hianyzo_runtime_mezo:{field_name}")
    warnings.extend(_runtime_field_warnings(normalized, row_index))

    generated_print_id = generate_print_id(normalized)
    if generated_print_id:
        normalized["generated_print_id"] = generated_print_id
    else:
        warnings.append(f"line={row_index} generated_print_id_nem_kepezheto")

    normalized["_source"] = "jsonl"
    normalized["_row_index"] = row_index
    normalized["_validation_issues"] = validate_row_mapping(normalized, row_index=row_index, sheet_name="EXPORT_RUNTIME")
    warnings.extend(normalized["_validation_issues"])
    return normalized, warnings


def load_card_rows_jsonl(fajl_utvonal):
    if not os.path.exists(fajl_utvonal):
        raise FileNotFoundError(fajl_utvonal)

    rows = []
    validation_issues = []
    seen_card_ids = {}

    with open(fajl_utvonal, "r", encoding="utf-8-sig") as handle:
        for row_index, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                raw_mapping = json.loads(line)
            except json.JSONDecodeError as exc:
                validation_issues.append(f"line={row_index} invalid_json:{exc.msg}")
                continue

            normalized, warnings = _normalize_runtime_row(raw_mapping, row_index)
            validation_issues.extend(warnings)
            if normalized is None:
                continue

            card_id = str(normalized.get("card_id") or "").strip()
            if not card_id:
                validation_issues.append(f"line={row_index} hianyzo_runtime_mezo:card_id")
            elif card_id in seen_card_ids:
                validation_issues.append(
                    f"line={row_index} duplicate_card_id:{card_id}:first_line={seen_card_ids[card_id]}"
                )
            else:
                seen_card_ids[card_id] = row_index

            rows.append(normalized)

    return rows, validation_issues


def load_card_rows_xlsx(fajl_utvonal):
    if not os.path.exists(fajl_utvonal):
        raise FileNotFoundError(fajl_utvonal)

    wb = load_workbook(fajl_utvonal, data_only=True)
    rows = []
    validation_issues = []

    for lap_nev in wb.sheetnames:
        sheet = wb[lap_nev]
        raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [_normalize_header_name(value) for value in raw_headers]

        if len(headers) != len(STANDARD_COLUMN_ORDER):
            validation_issues.append(
                f"sheet={lap_nev} helytelen_oszlopszam: vart={len(STANDARD_COLUMN_ORDER)} kapott={len(headers)}"
            )

        for expected in STANDARD_COLUMN_ORDER:
            if expected not in headers:
                validation_issues.append(f"sheet={lap_nev} hianyzo_oszlop:{expected}")

        for row_index, sor in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not sor or not sor[0]:
                continue

            first_cell = normalize_lookup_text(str(sor[0]).strip()) if sor[0] is not None else ""
            third_cell = normalize_lookup_text(str(sor[2]).strip()) if len(sor) > 2 and sor[2] is not None else ""
            if first_cell in {"kartya_nev", "kartya nev"} or third_cell == "birodalom":
                continue

            sor_mapping = _row_to_mapping(headers, sor)
            normalized = normalize_row_mapping(sor_mapping)
            normalized["_sheet"] = lap_nev
            normalized["_row_index"] = row_index
            normalized["_validation_issues"] = validate_row_mapping(normalized, row_index=row_index, sheet_name=lap_nev)
            validation_issues.extend(normalized["_validation_issues"])
            rows.append(normalized)

    return rows, validation_issues


def kartyak_betoltese_xlsx(fajl_utvonal):
    if not os.path.exists(fajl_utvonal):
        naplo.ir(f"HIBA: Nem talalhato a kartyafajl itt: {fajl_utvonal}")
        return []

    naplo.ir(f"Excel betoltese: {fajl_utvonal}")
    rows, validation_issues = load_card_rows_xlsx(fajl_utvonal)
    osszes_kartya = [Kartya(row) for row in rows]

    if validation_issues:
        naplo.ir(f"[VALIDATION] {len(validation_issues)} figyelmeztetes a cards.xlsx betoltes soran.")
        for issue in validation_issues[:20]:
            naplo.ir(f"[VALIDATION] {issue}")
        if len(validation_issues) > 20:
            naplo.ir(f"[VALIDATION] ... tovabbi {len(validation_issues) - 20} figyelmeztetes")

    naplo.ir(f"Sikeresen betoltve {len(osszes_kartya)} kartya.")
    return osszes_kartya


def kartyak_betoltese_jsonl(fajl_utvonal):
    if not os.path.exists(fajl_utvonal):
        naplo.ir(f"HIBA: Nem talalhato a kartyafajl itt: {fajl_utvonal}")
        return []

    naplo.ir(f"JSONL betoltese: {fajl_utvonal}")
    rows, validation_issues = load_card_rows_jsonl(fajl_utvonal)
    osszes_kartya = [Kartya(row) for row in rows]

    if validation_issues:
        naplo.ir(f"[VALIDATION] {len(validation_issues)} figyelmeztetes az EXPORT_RUNTIME.jsonl betoltes soran.")
        for issue in validation_issues[:20]:
            naplo.ir(f"[VALIDATION] {issue}")
        if len(validation_issues) > 20:
            naplo.ir(f"[VALIDATION] ... tovabbi {len(validation_issues) - 20} figyelmeztetes")

    naplo.ir(f"Sikeresen betoltve {len(osszes_kartya)} kartya.")
    return osszes_kartya


def kartyak_betoltese(fajl_utvonal):
    if str(fajl_utvonal).lower().endswith(".jsonl"):
        return kartyak_betoltese_jsonl(fajl_utvonal)
    return kartyak_betoltese_xlsx(fajl_utvonal)
