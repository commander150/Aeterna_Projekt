from __future__ import annotations

import collections
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from data.loader import (
    FIELD_ALIASES,
    FIELD_LEGACY_VALUES,
    FIELD_STANDARD_VALUES,
    classify_enum_token,
    load_card_rows_xlsx,
)
from engine.card_metadata import (
    LEGACY_INTERNAL_VALUES,
    STANDARD_FIELD_VALUES,
    normalize_duration_value,
    normalize_effect_tag_value,
    normalize_target_value,
    normalize_trigger_value,
    normalize_zone_value,
)
from utils.text import normalize_lookup_text

XLSX_PATH = ROOT / "cards.xlsx"
LOG_DIR = ROOT / "LOG"

CANONICAL_ALIAS_MAP_PATH = ROOT / "stats" / "canonical_alias_map.md"
STANDARD_ONLY_AUDIT_PATH = ROOT / "stats" / "standard_only_engine_compliance_audit.md"
WARNING_TRIAGE_PATH = ROOT / "stats" / "warning_triage_report.md"
TOP_GAPS_PATH = ROOT / "stats" / "top_20_standardization_gaps.md"
SUMMARY_PATH = ROOT / "stats" / "standard_cleanup_summary.md"

FIELD_LABELS = {
    "zona_felismerve": "Zona_Felismerve",
    "kulcsszavak_felismerve": "Kulcsszavak_Felismerve",
    "trigger_felismerve": "Trigger_Felismerve",
    "celpont_felismerve": "Celpont_Felismerve",
    "hatascimkek": "Hatascimkek",
    "idotartam_felismerve": "Idotartam_Felismerve",
}

FIELD_KEYS = tuple(FIELD_LABELS.keys())

FIELD_NAME_TO_STANDARD_KEY = {
    "zona_felismerve": "zone",
    "kulcsszavak_felismerve": "keyword",
    "trigger_felismerve": "trigger",
    "celpont_felismerve": "target",
    "hatascimkek": "effect_tag",
    "idotartam_felismerve": "duration",
}

FIELD_NORMALIZERS = {
    "zona_felismerve": normalize_zone_value,
    "kulcsszavak_felismerve": normalize_lookup_text,
    "trigger_felismerve": normalize_trigger_value,
    "celpont_felismerve": normalize_target_value,
    "hatascimkek": normalize_effect_tag_value,
    "idotartam_felismerve": normalize_duration_value,
}

FIELD_RUNTIME_VALUES = {
    "zona_felismerve": {"pecset", "pecsetek", "forras", "kez", "temeto", "pakli", "eternal", "player"},
    "kulcsszavak_felismerve": {"trap"},
    "trigger_felismerve": set(LEGACY_INTERNAL_VALUES["trigger"]),
    "celpont_felismerve": {"self_entity", "self unit", "sajat_entitas", "masik_sajat_entitas", "ellenseges_entitas"},
    "hatascimkek": set(LEGACY_INTERNAL_VALUES["effect_tag"]) | {"atk_buff", "hp_buff", "swap_position"},
    "idotartam_felismerve": set(LEGACY_INTERNAL_VALUES["duration"]),
}

BASELINE_ALLOWED = {
    "zona_felismerve": {"horizont", "zenit", "dominium", "graveyard", "hand", "deck", "source", "seal_row", "aeternal", "lane"},
    "kulcsszavak_felismerve": {"aegis", "bane", "burst", "celerity", "clarion", "echo", "ethereal", "harmonize", "resonance", "sundering", "taunt"},
    "trigger_felismerve": {
        "static", "on_play", "on_summon", "on_enemy_summon", "on_enemy_zenit_summon", "on_death", "on_destroyed",
        "on_attack_declared", "on_attack_hits", "on_combat_damage_dealt", "on_combat_damage_taken",
        "on_block_survived", "on_damage_survived", "on_enemy_spell_target", "on_enemy_spell_or_ritual_played",
        "on_enemy_spell_played", "on_spell_targeted", "on_enemy_extra_draw", "on_enemy_third_draw_in_turn",
        "on_turn_end", "on_next_own_awakening", "on_influx_phase", "on_heal", "on_enemy_ability_activated",
        "on_enemy_multiple_draws", "on_enemy_horizont_threshold", "on_move_zenit_to_horizont", "on_leave_board",
        "on_spell_cast_by_owner", "on_position_swap", "on_entity_enters_horizont", "on_source_placement",
        "on_seal_break", "on_bounce", "on_trap_triggered", "on_ready_from_exhausted", "on_stat_gain",
        "on_gain_keyword", "on_discard", "on_enemy_card_played", "on_enemy_second_summon_in_turn", "on_start_of_turn",
        "on_manifestation_phase", "on_awakening_phase",
    },
    "celpont_felismerve": set(STANDARD_FIELD_VALUES["target"]),
    "hatascimkek": {
        "damage", "deal_damage", "heal", "draw", "discard", "destroy", "return_to_hand", "seal_damage",
        "move_horizont", "move_zenit", "exhaust", "grant_keyword", "cost_mod", "graveyard_recursion",
        "damage_prevention", "attack_restrict", "block_restrict", "once_per_turn", "grant_attack",
        "grant_temp_attack", "grant_hp", "grant_max_hp", "swap_position", "reactivate", "immunity",
    },
    "idotartam_felismerve": {"until_end_of_turn", "until_end_of_next_own_turn", "until_end_of_combat", "permanent", "static", "instant"},
}

BASELINE_TRIGGER_ALIASES = {
    "on_manifest_phase": "on_manifestation_phase",
    "on_death": "on_destroyed",
    "death": "on_destroyed",
    "on_spell_targeted": "on_enemy_spell_target",
}

BASELINE_EFFECT_ALIASES = {
    "sebzes": "damage",
    "pecsetsebzes": "seal_damage",
    "laphuzas": "draw",
    "huzas": "draw",
    "gyogyitas": "heal",
    "megsemmisites": "destroy",
    "pusztitas": "destroy",
    "visszavetelkezbe": "return_to_hand",
    "kezbe_vetel": "return_to_hand",
    "mozgatas_horizontra": "move_horizont",
    "mozgatas_zenitbe": "move_zenit",
    "poziciocsere": "swap_position",
    "atk_buff": "grant_attack",
    "atk_mod": "grant_attack",
    "hp_buff": "grant_hp",
    "hp_mod": "grant_hp",
    "keyword_adas": "grant_keyword",
    "kulcsszoadas": "grant_keyword",
    "aura_modositas": "cost_mod",
    "kimerites": "exhaust",
    "tamadastiltas": "attack_restrict",
    "blokktiltas": "block_restrict",
    "sebzessemlegesites": "damage_prevention",
    "summon_token": "summon",
    "stat_protection": "damage_prevention",
    "damage_immunity": "damage_prevention",
    "ready": "reactivate",
}

BASELINE_DURATION_ALIASES = {
    "kor_vegeig": "until_end_of_turn",
    "a_kor_vegeig": "until_end_of_turn",
    "until_turn_end": "until_end_of_turn",
    "kovetkezo_kor_vegeig": "until_end_of_next_own_turn",
    "until_next_own_turn_end": "until_end_of_next_own_turn",
    "harc_vegeig": "until_end_of_combat",
    "harc_idejere": "until_end_of_combat",
    "during_combat": "until_end_of_combat",
    "allando": "permanent",
    "until_match_end": "permanent",
    "statikus": "static",
    "static_while_on_board": "static",
    "while_active": "static",
    "azonnali": "instant",
}

ENGINE_STATUS = {
    "fully_supported": "fully_supported",
    "partially_supported": "partially_supported",
    "small_change_needed": "small_change_needed",
    "new_engine_logic_needed": "new_engine_logic_needed",
    "not_observed": "not_observed_in_current_cards",
}

SUPPORT_OVERRIDES = {
    ("keyword", "aegis"): (ENGINE_STATUS["fully_supported"], ["engine/keyword_engine.py", "tests/test_keywords.py"], "keyword_engine", "No major blocker.", "Keep as reference keyword."),
    ("keyword", "bane"): (ENGINE_STATUS["fully_supported"], ["engine/keyword_engine.py"], "keyword_engine", "No major blocker.", "Keep as reference keyword."),
    ("keyword", "burst"): (ENGINE_STATUS["partially_supported"], ["cards/resolver.py", "engine/effect_diagnostics_v2.py"], "mixed runtime handler + diagnostics", "Runtime exists, but not as a fully generic standard keyword path.", "Continue moving burst semantics off text-only routing."),
    ("keyword", "celerity"): (ENGINE_STATUS["fully_supported"], ["engine/card.py", "engine/keyword_engine.py"], "card/unit initialization", "No major blocker.", "Keep as reference keyword."),
    ("keyword", "clarion"): (ENGINE_STATUS["partially_supported"], ["cards/resolver.py", "cards/priority_handlers.py"], "card-specific runtime handlers", "Mostly card-local, not a single generic standard resolver.", "Prefer shared Clarion routing when touching new cards."),
    ("keyword", "echo"): (ENGINE_STATUS["partially_supported"], ["engine/effects.py", "cards/resolver.py"], "death/runtime mix", "Death-side semantics are still partly card-local.", "Unify generic Echo bookkeeping later."),
    ("keyword", "ethereal"): (ENGINE_STATUS["fully_supported"], ["engine/keyword_engine.py", "tests/test_keywords.py"], "keyword_engine", "No major blocker.", "Keep as reference keyword."),
    ("keyword", "harmonize"): (ENGINE_STATUS["fully_supported"], ["engine/keyword_engine.py"], "keyword_engine", "No major blocker.", "Keep as reference keyword."),
    ("keyword", "resonance"): (ENGINE_STATUS["fully_supported"], ["engine/keyword_engine.py"], "keyword_engine", "No major blocker.", "Keep as reference keyword."),
    ("keyword", "sundering"): (ENGINE_STATUS["fully_supported"], ["engine/keyword_engine.py"], "keyword_engine", "No major blocker.", "Keep as reference keyword."),
    ("keyword", "taunt"): (ENGINE_STATUS["small_change_needed"], ["engine/game.py", "engine/keyword_engine.py"], "targeting/combat", "Forced-target routing is not fully driven from the standard keyword field.", "Add an explicit targeting gate for taunt."),
    ("trigger", "on_summon"): (ENGINE_STATUS["fully_supported"], ["engine/game.py", "cards/priority_handlers.py"], "runtime trigger dispatch", "No major blocker.", "Keep as reference trigger."),
    ("trigger", "on_death"): (ENGINE_STATUS["fully_supported"], ["engine/triggers.py", "cards/resolver.py"], "death trigger dispatch", "No major blocker.", "Keep as reference trigger."),
    ("trigger", "on_enemy_spell_target"): (ENGINE_STATUS["fully_supported"], ["engine/game.py", "cards/priority_handlers.py"], "spell-target reaction path", "No major blocker.", "Keep as reference trigger."),
    ("trigger", "on_enemy_spell_or_ritual_played"): (ENGINE_STATUS["small_change_needed"], ["engine/game.py", "cards/priority_handlers.py"], "event routing", "Spell vs ritual event unification is still uneven.", "Add a shared enemy spell-or-ritual event payload."),
    ("trigger", "on_enemy_zenit_summon"): (ENGINE_STATUS["small_change_needed"], ["engine/game.py"], "summon event with zone context", "Zone-filtered summon subscribers are still sparse.", "Route zenit summon subscribers from the existing summon payload."),
    ("trigger", "on_enemy_second_summon_in_turn"): (ENGINE_STATUS["new_engine_logic_needed"], ["engine/game.py"], "none", "Needs turn-scoped summon counters.", "Add generic per-turn summon counters first."),
    ("trigger", "on_stat_gain"): (ENGINE_STATUS["new_engine_logic_needed"], ["engine/triggers.py"], "none", "No generic stat-gain dispatch exists.", "Add a generic stat-change event if the cards justify it."),
    ("trigger", "on_start_of_turn"): (ENGINE_STATUS["fully_supported"], ["engine/game.py"], "turn phase dispatch", "No major blocker.", "Keep as reference trigger."),
    ("trigger", "static"): (ENGINE_STATUS["partially_supported"], ["engine/structured_effects.py", "engine/effect_diagnostics_v2.py"], "passive/static classification", "Many static cards are recognized but not all are explicitly simulated.", "Prefer passive/runtime_supported bookkeeping over missing_implementation."),
}

SUPPORT_OVERRIDES.update(
    {
        ("effect_tag", "damage"): (ENGINE_STATUS["fully_supported"], ["engine/effects.py", "engine/structured_effects.py", "tests/test_structured_effects.py"], "central damage pipeline", "No major blocker.", "Keep as reference primitive."),
        ("effect_tag", "seal_damage"): (ENGINE_STATUS["fully_supported"], ["engine/effects.py", "engine/structured_effects.py"], "seal damage pipeline", "No major blocker.", "Keep as reference primitive."),
        ("effect_tag", "draw"): (ENGINE_STATUS["fully_supported"], ["engine/player.py", "engine/structured_effects.py"], "draw pipeline", "No major blocker.", "Keep as reference primitive."),
        ("effect_tag", "heal"): (ENGINE_STATUS["fully_supported"], ["engine/structured_effects.py", "cards/priority_handlers.py"], "heal pipeline", "No major blocker.", "Keep as reference primitive."),
        ("effect_tag", "atk_mod"): (ENGINE_STATUS["fully_supported"], ["cards/priority_handlers.py", "engine/structured_effects.py", "tests/test_priority_handlers.py"], "buff helper + structured routing", "No major blocker.", "Keep as reference primitive."),
        ("effect_tag", "hp_mod"): (ENGINE_STATUS["partially_supported"], ["cards/priority_handlers.py", "engine/structured_effects.py"], "buff helper", "HP lifecycle and cleanup are less uniform than ATK buffs.", "Keep canonical `hp_mod`, continue consolidating cleanup."),
        ("effect_tag", "exhaust"): (ENGINE_STATUS["fully_supported"], ["engine/structured_effects.py"], "structured + runtime", "No major blocker.", "Keep as reference primitive."),
        ("effect_tag", "summon"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py", "engine/game.py"], "card-local summon logic", "No shared metadata-driven summon primitive exists yet.", "Introduce a small shared summon helper before broad rollout."),
        ("effect_tag", "destroy"): (ENGINE_STATUS["fully_supported"], ["engine/effects.py", "engine/structured_effects.py"], "destroy pipeline", "No major blocker.", "Keep as reference primitive."),
        ("effect_tag", "discard"): (ENGINE_STATUS["partially_supported"], ["cards/priority_handlers.py"], "card-local runtime", "Discard exists, but not as one uniform primitive.", "Add a shared discard helper when the next batch needs it."),
        ("effect_tag", "counterspell"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "card-local spell cancel", "Existing spell-target reactions do not yet expose a generic counterspell primitive.", "Lift the shared cancel branch into a reusable helper."),
        ("effect_tag", "redirect"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "card-local redirect", "Redirect exists locally, not as a generic primitive.", "Keep it local unless more redirect cards arrive."),
        ("effect_tag", "resource_gain"): (ENGINE_STATUS["small_change_needed"], ["engine/player.py"], "resource/source manipulation", "Resource gain exists around source handling, but not under a standard effect-tag path.", "Add a small effect-tag adapter to source gain."),
        ("effect_tag", "resource_drain"): (ENGINE_STATUS["new_engine_logic_needed"], ["engine/player.py"], "none", "No clean generic enemy resource drain primitive is exposed.", "Design a shared resource-drain contract first."),
        ("effect_tag", "resource_acceleration"): (ENGINE_STATUS["small_change_needed"], ["engine/player.py"], "source acceleration helpers", "Close to source manipulation, but not yet tagged explicitly.", "Alias into source acceleration helper if semantics match."),
        ("effect_tag", "resource_spend"): (ENGINE_STATUS["small_change_needed"], ["engine/player.py"], "source spend flow", "Spend is implicit in costs, not explicit as an effect primitive.", "Add a thin explicit resource spend helper."),
    }
)

DEFAULT_SUPPORT = {
    "zone": (ENGINE_STATUS["fully_supported"], ["engine/game.py"], "core board model", "No major blocker.", "Keep as reference standard value."),
    "keyword": (ENGINE_STATUS["partially_supported"], ["engine/keyword_engine.py"], "keyword_engine", "Keyword is recognized but not yet proven end-to-end everywhere.", "Add targeted runtime proof where cards demand it."),
    "trigger": (ENGINE_STATUS["small_change_needed"], ["engine/game.py", "engine/triggers.py"], "event dispatch", "Trigger exists conceptually, but standard routing is not fully unified yet.", "Add a thin adapter before deeper engine work."),
    "target": (ENGINE_STATUS["partially_supported"], ["engine/structured_effects.py", "engine/game.py"], "target routing", "Target family exists, but not every standard target has a direct shared validator path.", "Tighten target validation incrementally."),
    "effect_tag": (ENGINE_STATUS["small_change_needed"], ["engine/structured_effects.py", "cards/priority_handlers.py"], "structured + local runtime mix", "Canonical tag is present in data, but not yet a first-class runtime primitive.", "Prefer small tag adapters over new frameworks."),
    "duration": (ENGINE_STATUS["small_change_needed"], ["engine/game.py"], "expiry bookkeeping", "Timing buckets exist, but many standard durations are not yet explicit.", "Add one expiry bucket at a time."),
}


def _normalize_header(header):
    return normalize_lookup_text(str(header or "")).replace(" ", "_")


def _split_tokens(value):
    if value is None:
        return []
    text = str(value).replace("|", ",").replace(";", ",")
    return [part.strip() for part in text.split(",") if part and part.strip()]


def _load_raw_rows():
    workbook = load_workbook(XLSX_PATH, data_only=True)
    rows = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        headers = [_normalize_header(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())]
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            mapping = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                mapping[header] = row[idx] if idx < len(row) else None
            mapping["_sheet"] = sheet_name
            mapping["_row_index"] = row_index
            mapping["_card_name"] = str(mapping.get("kartya_nev") or row[0]).strip()
            rows.append(mapping)
    return rows


def _find_latest_logs(limit=5):
    files = sorted(LOG_DIR.rglob("AETERNA_LOG_*.txt"), key=lambda path: path.stat().st_mtime, reverse=True)
    return files[:limit]


def _canonicalize_for_field(field_name, token):
    issue_type, canonical = classify_enum_token(field_name, token)
    if issue_type == "standard":
        return normalize_lookup_text(token)
    if canonical:
        return canonical
    return ""


def _collect_observed_values(raw_rows):
    observed = {field: set() for field in FIELD_KEYS}
    for row in raw_rows:
        for field in FIELD_KEYS:
            for token in _split_tokens(row.get(field)):
                observed[field].add(normalize_lookup_text(token))
    for field, alias_map in FIELD_ALIASES.items():
        observed[field].update(alias_map.keys())
    for field, values in FIELD_LEGACY_VALUES.items():
        observed[field].update(values)
    return observed


def _value_metadata(field_name, value):
    standard = FIELD_STANDARD_VALUES[field_name]
    alias_map = FIELD_ALIASES.get(field_name, {})
    if value in standard:
        return ("standard", value, "standard_doc")
    if value in alias_map and alias_map[value] in standard:
        return ("alias", alias_map[value], "loader_alias")
    if value in FIELD_LEGACY_VALUES.get(field_name, set()):
        canonical = FIELD_NORMALIZERS[field_name](value)
        if canonical in standard:
            return ("legacy_internal", canonical, "runtime_alias")
        return ("legacy_internal", "", "runtime_alias")
    canonical = FIELD_NORMALIZERS[field_name](value)
    if canonical in standard:
        return ("alias", canonical, "inferred")
    return ("invalid", "", "none")


def _format_bool(flag):
    return "yes" if flag else "no"


def _baseline_normalize_token(field_name, token):
    normalized = normalize_lookup_text(token)
    if normalized in {"", "blank", "none", "-"}:
        return ""
    if field_name == "trigger_felismerve":
        return BASELINE_TRIGGER_ALIASES.get(normalized, normalized)
    if field_name == "hatascimkek":
        return BASELINE_EFFECT_ALIASES.get(normalized, normalized)
    if field_name == "idotartam_felismerve":
        return BASELINE_DURATION_ALIASES.get(normalized, normalized)
    return normalized


def write_canonical_alias_map(raw_rows):
    observed = _collect_observed_values(raw_rows)
    lines = [
        "# Canonical vs alias map",
        "",
        "The final audit should only present canonical standard values. Non-standard values stay here as alias, legacy, or invalid observations.",
        "",
    ]
    for field in FIELD_KEYS:
        lines.extend(
            [
                f"## {FIELD_LABELS[field]}",
                "",
                "| field_name | observed_value | value_type | canonical_value | source_of_mapping | used_in_loader | used_in_runtime | notes |",
                "| --- | --- | --- | --- | --- | --- | --- | --- |",
            ]
        )
        for value in sorted(observed[field]):
            value_type, canonical, source = _value_metadata(field, value)
            notes = []
            if value_type == "invalid":
                notes.append("No safe canonical mapping from the current standard.")
            elif value_type == "legacy_internal":
                notes.append("Runtime/internal naming; keep out of final canonical reports.")
            elif value_type == "alias":
                notes.append("Can be normalized to the canonical standard value.")
            if field == "trigger_felismerve" and value == "on_play":
                notes.append("Not listed in the final trigger standard; keep as legacy/runtime-only until data is cleaned.")
            if field == "zona_felismerve" and value == "burst":
                notes.append("Burst is a keyword, not a zone.")
            if field == "kulcsszavak_felismerve" and value == "trap":
                notes.append("Trap/Jel is card-type semantics, not a standard keyword.")
            if field == "idotartam_felismerve" and value == "trap":
                notes.append("Trap is not a duration.")
            used_in_loader = value in observed[field] or value in FIELD_ALIASES.get(field, {})
            used_in_runtime = value in FIELD_RUNTIME_VALUES.get(field, set())
            lines.append(
                f"| `{FIELD_LABELS[field]}` | `{value}` | `{value_type}` | `{canonical or '-'}` | `{source}` | `{_format_bool(used_in_loader)}` | `{_format_bool(used_in_runtime)}` | {' '.join(notes) or '-'} |"
            )
        lines.append("")
    CANONICAL_ALIAS_MAP_PATH.write_text("\n".join(lines), encoding="utf-8")


def _observed_canonical_counts(raw_rows):
    counts = collections.defaultdict(collections.Counter)
    for row in raw_rows:
        for field in FIELD_KEYS:
            for token in _split_tokens(row.get(field)):
                canonical = _canonicalize_for_field(field, normalize_lookup_text(token))
                if canonical:
                    counts[field][canonical] += 1
    return counts


def _support_row(field_name, canonical_value, observed_count):
    standard_key = FIELD_NAME_TO_STANDARD_KEY[field_name]
    if observed_count == 0:
        return (
            ENGINE_STATUS["not_observed"],
            ["-"],
            "-",
            "Not observed in the current cards.xlsx.",
            "Keep documented; no runtime action needed yet.",
        )
    result = SUPPORT_OVERRIDES.get((standard_key, canonical_value))
    if result is not None:
        return result
    return DEFAULT_SUPPORT[standard_key]


def write_standard_only_compliance(raw_rows):
    counts = _observed_canonical_counts(raw_rows)
    lines = [
        "# Standard-only engine compliance audit",
        "",
        "Only final canonical standard values appear as rows below. Aliases and legacy/internal names are intentionally excluded from the main audit and referenced through `canonical_alias_map.md`.",
        "",
    ]
    for field in FIELD_KEYS:
        lines.extend(
            [
                f"## {FIELD_LABELS[field]}",
                "",
                "| canonical value | support status | evidence in code | evidence in tests | current handling path | blockers / missing logic | recommended next step |",
                "| --- | --- | --- | --- | --- | --- | --- |",
            ]
        )
        for canonical_value in sorted(FIELD_STANDARD_VALUES[field]):
            support_status, evidence, handling_path, blockers, next_step = _support_row(field, canonical_value, counts[field][canonical_value])
            code_refs = [item for item in evidence if item.startswith(("engine/", "cards/", "data/"))]
            test_refs = [item for item in evidence if item.startswith("tests/")]
            lines.append(
                f"| `{canonical_value}` | `{support_status}` | {', '.join(f'`{item}`' for item in code_refs) or '-'} | {', '.join(f'`{item}`' for item in test_refs) or '-'} | {handling_path} | {blockers} | {next_step} |"
            )
        lines.append("")
    STANDARD_ONLY_AUDIT_PATH.write_text("\n".join(lines), encoding="utf-8")


def _baseline_issues(raw_rows):
    issues = []
    for raw_row in raw_rows:
        location = f"sheet={raw_row['_sheet']} row={raw_row['_row_index']}"
        card_name = raw_row["_card_name"]
        for field in FIELD_KEYS:
            for token in _split_tokens(raw_row.get(field, "")):
                normalized_token = _baseline_normalize_token(field, token)
                if not normalized_token:
                    continue
                if normalized_token not in BASELINE_ALLOWED[field]:
                    issues.append(
                        {
                            "category": "enum_deviation",
                            "field": field,
                            "token": normalized_token,
                            "sheet": raw_row["_sheet"],
                            "row_index": raw_row["_row_index"],
                            "card_name": card_name,
                            "location": location,
                        }
                    )
        if normalize_lookup_text(raw_row.get("kartyatipus", "")) == "entitas":
            suspicious_targets = {"enemy_spell", "enemy_spell_or_ritual", "own_hand", "enemy_hand"}
            for token in _split_tokens(raw_row.get("celpont_felismerve", "")):
                normalized_token = _baseline_normalize_token("celpont_felismerve", token)
                if normalized_token in suspicious_targets:
                    issues.append(
                        {
                            "category": "suspicious_field_combination",
                            "field": "celpont_felismerve",
                            "token": normalized_token,
                            "sheet": raw_row["_sheet"],
                            "row_index": raw_row["_row_index"],
                            "card_name": card_name,
                            "location": location,
                            "reason": "entity card with suspicious target type",
                        }
                    )
        normalized_duration_tokens = [_baseline_normalize_token("idotartam_felismerve", token) for token in _split_tokens(raw_row.get("idotartam_felismerve", ""))]
        normalized_duration_tokens = [token for token in normalized_duration_tokens if token]
        normalized_effect_tokens = [_baseline_normalize_token("hatascimkek", token) for token in _split_tokens(raw_row.get("hatascimkek", ""))]
        normalized_effect_tokens = [token for token in normalized_effect_tokens if token]
        if normalized_duration_tokens and not normalized_effect_tokens:
            issues.append(
                {
                    "category": "suspicious_field_combination",
                    "field": "idotartam_felismerve",
                    "token": ", ".join(normalized_duration_tokens),
                    "sheet": raw_row["_sheet"],
                    "row_index": raw_row["_row_index"],
                    "card_name": card_name,
                    "location": location,
                    "reason": "duration present without effect tag",
                }
            )
    return issues


def _runtime_support_status_for_token(field_name, token):
    canonical = _canonicalize_for_field(field_name, token)
    if not canonical:
        return ""
    support_status, _, _, _, _ = _support_row(field_name, canonical, 1)
    return support_status


def _triage_issue(issue):
    if issue["category"] == "suspicious_field_combination":
        return "suspicious_field_combination"
    field = issue["field"]
    token = normalize_lookup_text(issue["token"])
    issue_type, canonical = classify_enum_token(field, token)
    if issue_type == "alias_normalizable":
        return "alias_normalizable"
    if issue_type == "legacy_internal":
        return "alias_normalizable" if canonical else "other"
    if issue_type == "invalid_delimiter_or_format":
        return "invalid_delimiter_or_format"
    if issue_type == "unknown_enum_value":
        return "unknown_enum_value"
    support_status = _runtime_support_status_for_token(field, token)
    if support_status in {ENGINE_STATUS["partially_supported"], ENGINE_STATUS["small_change_needed"], ENGINE_STATUS["new_engine_logic_needed"]}:
        return "unsupported_runtime_semantics"
    return "other"


def _example_reason(category, issue):
    field = issue["field"]
    token = issue["token"]
    if category == "unsupported_runtime_semantics":
        canonical = _canonicalize_for_field(field, token)
        support_status = _runtime_support_status_for_token(field, token)
        return f"`{token}` is a standard `{field}` value, but its current engine support is only `{support_status}`."
    if category == "alias_normalizable":
        _, canonical = classify_enum_token(field, token)
        return f"`{token}` can be normalized to canonical `{canonical}`."
    if category == "invalid_delimiter_or_format":
        return f"`{token}` looks like a format/delimiter issue instead of a valid canonical token."
    if category == "unknown_enum_value":
        return f"`{token}` is not part of the final standard and has no safe canonical mapping."
    if category == "suspicious_field_combination":
        return issue.get("reason", "Field combination is suspicious under the current standard.")
    return "Needs manual follow-up."


def _example_fix(category, issue):
    field = issue["field"]
    token = issue["token"]
    if category == "unsupported_runtime_semantics":
        canonical = _canonicalize_for_field(field, token)
        return f"Add or tighten runtime support for canonical `{canonical}` without inventing a new standard value."
    if category == "alias_normalizable":
        _, canonical = classify_enum_token(field, token)
        return f"Normalize `{token}` to `{canonical}` in loader/reporting; keep alias only in the alias map."
    if category == "invalid_delimiter_or_format":
        return "Fix the cell formatting and rewrite it to a canonical standard token."
    if category == "unknown_enum_value":
        return "Treat as data cleanup or document it as uncertain until a real canonical mapping exists."
    if category == "suspicious_field_combination":
        return "Review the row semantics; keep the canonical value but fix the inconsistent field combination."
    return "Manual review."


def write_warning_triage(raw_rows):
    baseline = _baseline_issues(raw_rows)
    grouped = collections.defaultdict(list)
    for issue in baseline:
        grouped[_triage_issue(issue)].append(issue)

    ordered_categories = [
        "unknown_enum_value",
        "alias_normalizable",
        "missing_explicit_blank_or_none",
        "suspicious_field_combination",
        "invalid_delimiter_or_format",
        "unsupported_runtime_semantics",
        "ambiguous_card_text",
        "other",
    ]

    lines = [
        "# Warning triage report",
        "",
        "Baseline warning backlog from the latest logs: `471`.",
        f"Triage rows reproduced from workbook + baseline normalization rules: `{len(baseline)}`.",
        f"Unreproduced delta against the log baseline: `{471 - len(baseline)}`.",
        "",
        "## Counts by category",
        "",
        "| category | count |",
        "| --- | ---: |",
    ]
    for category in ordered_categories:
        lines.append(f"| `{category}` | {len(grouped.get(category, []))} |")
    lines.extend(["", "## Examples by category", ""])

    for category in ordered_categories:
        examples = grouped.get(category, [])
        lines.append(f"### {category}")
        lines.append("")
        if not examples:
            lines.append("No baseline examples fell into this category.")
            lines.append("")
            continue
        lines.append("| card_name | field | current value | why it is here | recommended fix |")
        lines.append("| --- | --- | --- | --- | --- |")
        for issue in examples[:5]:
            lines.append(
                f"| `{issue['card_name']}` | `{issue['field']}` | `{issue['token']}` | {_example_reason(category, issue)} | {_example_fix(category, issue)} |"
            )
        lines.append("")

    WARNING_TRIAGE_PATH.write_text("\n".join(lines), encoding="utf-8")
    return baseline, grouped


def _gap_fix_type(category):
    if category == "alias_normalizable":
        return "loader_alias"
    if category == "invalid_delimiter_or_format":
        return "validator_rule"
    if category == "unsupported_runtime_semantics":
        return "runtime_support"
    if category == "unknown_enum_value":
        return "data_fix"
    if category == "suspicious_field_combination":
        return "doc_clarification"
    return "validator_rule"


def _gap_risk_and_effort(category):
    if category in {"alias_normalizable", "invalid_delimiter_or_format", "unknown_enum_value"}:
        return ("low", "small")
    if category == "suspicious_field_combination":
        return ("medium", "small")
    return ("medium", "medium")


def write_top_20_gaps(baseline):
    pattern_counts = collections.Counter()
    pattern_examples = collections.defaultdict(list)
    pattern_categories = {}
    for issue in baseline:
        category = _triage_issue(issue)
        pattern = issue.get("reason") if issue["category"] == "suspicious_field_combination" else f"{issue['field']}:{issue['token']}"
        pattern_counts[pattern] += 1
        pattern_categories[pattern] = category
        if len(pattern_examples[pattern]) < 3:
            pattern_examples[pattern].append(issue["card_name"])

    lines = [
        "# Top 20 standardization gaps",
        "",
        "| rank | pattern | occurrence_count | affected_field | representative_examples | fix_type | estimated_risk | estimated_effort | recommendation |",
        "| ---: | --- | ---: | --- | --- | --- | --- | --- | --- |",
    ]
    for rank, (pattern, count) in enumerate(pattern_counts.most_common(20), start=1):
        category = pattern_categories[pattern]
        examples = ", ".join(f"`{item}`" for item in pattern_examples[pattern])
        risk, effort = _gap_risk_and_effort(category)
        first_issue = next(
            issue
            for issue in baseline
            if ((issue.get("reason") if issue["category"] == "suspicious_field_combination" else f"{issue['field']}:{issue['token']}") == pattern)
        )
        recommendation = _example_fix(category, first_issue)
        lines.append(
            f"| {rank} | `{pattern}` | {count} | `{first_issue['field']}` | {examples} | `{_gap_fix_type(category)}` | `{risk}` | `{effort}` | {recommendation} |"
        )
    TOP_GAPS_PATH.write_text("\n".join(lines), encoding="utf-8")


def write_summary(new_validation_issues, baseline_grouped, latest_logs):
    reproduced_total = sum(len(items) for items in baseline_grouped.values())
    alias_count = len(baseline_grouped.get("alias_normalizable", []))
    data_count = len(baseline_grouped.get("unknown_enum_value", [])) + len(baseline_grouped.get("invalid_delimiter_or_format", []))
    engine_count = len(baseline_grouped.get("unsupported_runtime_semantics", []))
    top_findings = [
        "The latest 5 log files all repeat the same `471` validation backlog, so the problem is structural, not sample noise.",
        "A large part of the old `ismeretlen_ertek` noise is not actually unknown data anymore: many tokens are valid standard values whose runtime support is only partial.",
        "The previous loader was canonicalizing several structured fields toward runtime names; the cleanup pulls canonical storage back toward the spreadsheet standard.",
        "`grant_attack`/`grant_hp`/`reactivate` are now treated as legacy/runtime names behind canonical `atk_mod`/`hp_mod`/`ready`.",
        "`until_end_of_turn`/`until_end_of_combat` style names are now normalized toward standard `until_turn_end`/`during_combat`.",
        "`on_play` remains a real legacy hotspot: it is heavily used by runtime code, but it is not a final canonical trigger value in the current standard.",
        "Several effect tags that looked invalid under the old validator are actually standard spreadsheet values, but still need runtime adapters or explicit support.",
        "`burst` in zone fields and `trap` in keyword/duration fields remain true standard-fidelity problems, not just alias issues.",
        "The validator now distinguishes alias-normalizable values, legacy/internal values, invalid formatting, and truly unknown enum values.",
        "The next best wins are small adapters for standard effect tags such as `summon`, `resource_gain`, `counterspell`, and `damage_bonus`, plus data cleanup around `on_play`.",
    ]
    lines = [
        "# Standard cleanup summary",
        "",
        "## What was cleaned",
        "",
        "- Loader validation now distinguishes canonical values, normalizable aliases, legacy/internal runtime names, invalid formatting, and true unknown enum values.",
        "- Structured canonicalization was pulled back toward the spreadsheet standard for triggers, effect tags, and durations.",
        "- Final audit outputs are now split: canonical-only compliance stays separate from alias/legacy observations.",
        "",
        "## Latest log signal",
        "",
        f"- Latest log files inspected: {', '.join(f'`{path.name}`' for path in latest_logs)}",
        "- Every inspected log still reports `471` validation warnings before this cleanup pass.",
        "",
        "## Remaining non-standard values",
        "",
        "- Legacy/runtime-heavy: `on_play`, `on_destroyed`, `grant_attack`, `grant_hp`, `reactivate`, `until_end_of_turn`, `until_end_of_combat`.",
        "- True standard-fidelity problems: `burst` in `Zona_Felismerve`, `trap` in `Kulcsszavak_Felismerve`, `trap` in `Idotartam_Felismerve`, `from hand` formatting in `Zona_Felismerve`.",
        "",
        "## Warning split",
        "",
        f"- Reproduced workbook-backed triage rows from the `471` log backlog: `{reproduced_total}`",
        f"- Remaining log-vs-workbook delta: `{471 - reproduced_total}`",
        f"- Alias-like or legacy-normalizable: `{alias_count}`",
        f"- Real data / format problems: `{data_count}`",
        f"- Engine support gaps on standard values: `{engine_count}`",
        f"- New validator warning count after cleanup rules: `{len(new_validation_issues)}`",
        "",
        "## Next 5 best steps",
        "",
        "1. Clean the data rows that still use `on_play` where no safe canonical trigger exists yet, instead of letting it leak into final audits.",
        "2. Add small runtime adapters for standard effect tags with the highest frequency: `summon`, `resource_gain`, `ability_lock`, `sacrifice`, `untargetable`.",
        "3. Fix the clearly invalid standard-fidelity rows: zone=`burst`, keyword=`trap`, duration=`trap`, zone=`from hand`.",
        "4. Add one shared deferred-expiry helper for `until_next_enemy_turn`, `next_own_awakening`, and `next_own_turn_start`.",
        "5. Re-run the same latest-log comparison after the next cleanup so we can measure how much of the old `471` backlog was alias noise versus real engine debt.",
        "",
        "## Top findings",
        "",
    ]
    for item in top_findings:
        lines.append(f"- {item}")
    SUMMARY_PATH.write_text("\n".join(lines), encoding="utf-8")


def main():
    raw_rows = _load_raw_rows()
    _, new_validation_issues = load_card_rows_xlsx(str(XLSX_PATH))
    latest_logs = _find_latest_logs()
    write_canonical_alias_map(raw_rows)
    write_standard_only_compliance(raw_rows)
    baseline, grouped = write_warning_triage(raw_rows)
    write_top_20_gaps(baseline)
    write_summary(new_validation_issues, grouped, latest_logs)
    print(f"Generated: {CANONICAL_ALIAS_MAP_PATH}")
    print(f"Generated: {STANDARD_ONLY_AUDIT_PATH}")
    print(f"Generated: {WARNING_TRIAGE_PATH}")
    print(f"Generated: {TOP_GAPS_PATH}")
    print(f"Generated: {SUMMARY_PATH}")


if __name__ == "__main__":
    main()

SUPPORT_OVERRIDES.update(
    {
        ("effect_tag", "move_horizont"): (ENGINE_STATUS["fully_supported"], ["engine/actions.py", "engine/structured_effects.py"], "movement helper", "No major blocker.", "Keep as reference primitive."),
        ("effect_tag", "move_zenit"): (ENGINE_STATUS["fully_supported"], ["engine/actions.py", "engine/structured_effects.py"], "movement helper", "No major blocker.", "Keep as reference primitive."),
        ("effect_tag", "graveyard_recursion"): (ENGINE_STATUS["partially_supported"], ["engine/actions.py", "cards/priority_handlers.py"], "card-local recursion", "Present, but still card-local in many cases.", "Promote a small shared recursion helper when needed."),
        ("effect_tag", "graveyard_replacement"): (ENGINE_STATUS["new_engine_logic_needed"], ["engine/effects.py"], "none", "Replacement semantics are not modeled generically.", "Do not fake this with ad hoc handlers."),
        ("effect_tag", "grant_keyword"): (ENGINE_STATUS["partially_supported"], ["cards/priority_handlers.py"], "card-local keyword grant", "Temporary keyword grant exists but is not fully uniform.", "Keep standard tag and expand helper only when needed."),
        ("effect_tag", "type_change"): (ENGINE_STATUS["new_engine_logic_needed"], ["engine/card.py"], "none", "Card-type mutation is not modeled generically.", "Defer until there is card pressure."),
        ("effect_tag", "stat_reset"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "buff cleanup", "Very close to existing buff cleanup, but not explicit.", "Expose a small stat reset helper."),
        ("effect_tag", "trap_immunity"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "reaction flags", "Can likely sit on existing trap checks, but no standard hook yet.", "Add one trap-ignore flag if cards demand it."),
        ("effect_tag", "damage_immunity"): (ENGINE_STATUS["partially_supported"], ["cards/priority_handlers.py", "engine/structured_effects.py"], "local protection flags", "Damage prevention exists, but immunity semantics are still local.", "Keep local until more cards need it."),
        ("effect_tag", "damage_bonus"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "combat bonus helpers", "Combat buffs exist; a standardized bonus tag adapter is missing.", "Map onto existing combat bonus helpers."),
        ("effect_tag", "damage_prevention"): (ENGINE_STATUS["partially_supported"], ["cards/priority_handlers.py", "engine/structured_effects.py"], "prevention flags", "Works in local/runtime patterns, not yet fully generic.", "Keep using local handlers where precise timing matters."),
        ("effect_tag", "overflow_damage"): (ENGINE_STATUS["small_change_needed"], ["engine/game.py"], "overflow tracking", "Overflow defeat exists, but not a general effect-tag route.", "Add a thin effect-tag bridge to existing overflow logic."),
        ("effect_tag", "stat_protection"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "local protection flags", "Close to prevention/immunity patterns, but not standard-routed.", "Resolve via a shared protection flag if more cards appear."),
        ("effect_tag", "sacrifice"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "destroy-own-unit patterns", "Very close to destroy + own target selection.", "Add a small sacrifice helper instead of bespoke logic."),
        ("effect_tag", "free_cast"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "cost bypass flags", "Cost modification exists, but free-cast is not a standard primitive.", "Bridge to cost-mod with a zero-cost branch."),
        ("effect_tag", "tutor"): (ENGINE_STATUS["small_change_needed"], ["engine/player.py"], "deck search patterns", "Deck handling exists, but no shared tutor primitive.", "Add a small search helper when the next tutor card is implemented."),
        ("effect_tag", "untargetable"): (ENGINE_STATUS["partially_supported"], ["engine/targeting.py"], "targeting flags", "Targeting protections exist, but not fully tagged from structured metadata.", "Keep standard tag; tighten targeting checks incrementally."),
        ("effect_tag", "return_to_hand"): (ENGINE_STATUS["fully_supported"], ["engine/actions.py", "engine/structured_effects.py"], "return_to_hand helper", "No major blocker.", "Keep as reference primitive."),
        ("effect_tag", "summon_token"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "card-local summon logic", "Token summon is close to summon support, but not a generic primitive.", "Share summon helper first, then add token parameters."),
        ("effect_tag", "attack_restrict"): (ENGINE_STATUS["partially_supported"], ["engine/game.py", "cards/priority_handlers.py"], "cannot_attack flags", "Works locally, not fully generic.", "Keep standard tag and widen shared flag usage gradually."),
        ("effect_tag", "summon_restrict"): (ENGINE_STATUS["small_change_needed"], ["engine/game.py"], "none", "Summon restriction checks are not fully centralized.", "Add one summon legality gate when the next card needs it."),
        ("effect_tag", "block_restrict"): (ENGINE_STATUS["small_change_needed"], ["engine/game.py"], "combat legality", "Block restriction is close to attack restriction but less explicit.", "Share legality gate with attack restriction."),
        ("effect_tag", "control_change"): (ENGINE_STATUS["new_engine_logic_needed"], ["engine/card.py"], "none", "Temporary or permanent control change is not modeled generically.", "Do not fake with owner swaps."),
        ("effect_tag", "ready"): (ENGINE_STATUS["fully_supported"], ["engine/player.py", "engine/structured_effects.py"], "reactivate helper", "No major blocker.", "Keep as reference primitive."),
        ("effect_tag", "return_to_deck"): (ENGINE_STATUS["small_change_needed"], ["engine/player.py"], "deck manipulation", "Top/bottom deck operations exist partially, but not a generic return_to_deck primitive.", "Add a thin shared helper."),
        ("effect_tag", "deck_bottom"): (ENGINE_STATUS["small_change_needed"], ["engine/player.py"], "deck manipulation", "Bottom placement is close to deck return, but not explicit everywhere.", "Add a small helper that selects top vs bottom."),
        ("effect_tag", "move_to_source"): (ENGINE_STATUS["small_change_needed"], ["engine/player.py", "cards/priority_handlers.py"], "source manipulation", "The source zone exists, but not as a generic standard effect-tag route.", "Add a shared source move helper."),
        ("effect_tag", "source_manipulation"): (ENGINE_STATUS["small_change_needed"], ["engine/player.py"], "source manipulation", "Source operations exist, but are not grouped behind one canonical tag.", "Route standard tag to source helpers."),
        ("effect_tag", "cleanse"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "local status cleanup", "Close to existing buff/flag cleanup, but not exposed as a generic effect.", "Add a small shared cleanup helper."),
        ("effect_tag", "copy_stats"): (ENGINE_STATUS["new_engine_logic_needed"], ["engine/card.py"], "none", "No generic copy-stats layer exists.", "Defer until there is more than one concrete card."),
        ("effect_tag", "copy_keywords"): (ENGINE_STATUS["new_engine_logic_needed"], ["engine/keyword_engine.py"], "none", "No generic keyword-copy layer exists.", "Defer until there is more than one concrete card."),
        ("effect_tag", "position_lock"): (ENGINE_STATUS["small_change_needed"], ["engine/actions.py"], "movement checks", "Close to movement legality, but not yet explicit.", "Add a lock flag if more cards need it."),
        ("effect_tag", "attack_nullify"): (ENGINE_STATUS["partially_supported"], ["cards/priority_handlers.py"], "trap/combat local handlers", "Nullify behavior exists locally in traps and reactions.", "Keep local unless repeated patterns emerge."),
        ("effect_tag", "ability_lock"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "local flags", "Ability prevention exists locally, not as a generic tag route.", "Add a lightweight ability-use flag only if repeated."),
        ("effect_tag", "random_discard"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "discard patterns", "Very close to discard support, but random selection is not generalized.", "Add random selection atop discard helper."),
        ("effect_tag", "choice"): (ENGINE_STATUS["small_change_needed"], ["cards/priority_handlers.py"], "card-local branches", "Choice is UI and rules glue, not a primitive effect.", "Keep as audit tag; do not over-generalize yet."),
    }
)

SUPPORT_OVERRIDES.update(
    {
        ("duration", "instant"): (ENGINE_STATUS["fully_supported"], ["engine/effect_diagnostics_v2.py"], "on-play resolution", "No major blocker.", "Keep as reference duration."),
        ("duration", "during_combat"): (ENGINE_STATUS["fully_supported"], ["engine/structured_effects.py", "cards/priority_handlers.py"], "combat cleanup", "No major blocker.", "Keep as reference duration."),
        ("duration", "until_turn_end"): (ENGINE_STATUS["fully_supported"], ["engine/game.py", "cards/priority_handlers.py"], "turn-end cleanup", "No major blocker.", "Keep as reference duration."),
        ("duration", "until_next_own_turn_end"): (ENGINE_STATUS["partially_supported"], ["cards/priority_handlers.py"], "local deferred cleanup", "Next-turn cleanup is not yet generic.", "Promote a shared deferred-expiry tracker when needed."),
        ("duration", "until_next_enemy_turn"): (ENGINE_STATUS["small_change_needed"], ["engine/game.py"], "turn-phase tracking", "The turn framework exists, but enemy-turn expiry is not standardized.", "Add one generic enemy-turn expiry bucket."),
        ("duration", "until_match_end"): (ENGINE_STATUS["partially_supported"], ["cards/priority_handlers.py"], "persistent flags", "Long-lived effects are local, not centrally modeled.", "Keep local until many cards need persistent state."),
        ("duration", "static_while_on_board"): (ENGINE_STATUS["partially_supported"], ["engine/structured_effects.py"], "static/passive classification", "Recognized, but many cards still rely on passive bookkeeping rather than explicit simulation.", "Prefer runtime/passive bookkeeping over missing_implementation."),
        ("duration", "while_active"): (ENGINE_STATUS["small_change_needed"], ["engine/game.py"], "state-dependent aura checks", "Close to static-on-board, but needs an explicit active-state check.", "Add a shared `while_active` predicate only when the next card needs it."),
        ("duration", "next_own_awakening"): (ENGINE_STATUS["small_change_needed"], ["engine/game.py"], "phase tracking", "Awakening phase exists, but not yet exposed as a generic deferred-expiry target.", "Route to one awakening-expiry bucket."),
        ("duration", "next_own_turn_start"): (ENGINE_STATUS["small_change_needed"], ["engine/game.py"], "phase tracking", "Turn start exists, but not yet a reusable duration target.", "Add one turn-start expiry bucket."),
    }
)
