"""Standalone XLSX worksheet exporter for JSONL and TSV output."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook


PROGRAM_DIR = Path(__file__).resolve().parent
SOURCE_DIR = PROGRAM_DIR / "source"
DEFAULT_OUTPUT_DIR = PROGRAM_DIR / "exports"
LOOKUPS_FIELDS = (
    "Lookup_Group",
    "Value",
    "Label_HU",
    "Status",
    "Canonical_Value",
    "Used_For",
    "Sort_Order",
    "Source",
    "Notes",
)


class ExportError(Exception):
    """An expected error that can be shown directly to the user."""


@dataclass(frozen=True)
class ExportProfile:
    name: str
    sheet_name: str | None
    output_filename: str | None
    output_format: str
    required_fields: tuple[str, ...] = ()
    number_fields: tuple[str, ...] = ()
    output_fields: tuple[str, ...] = ()
    skip_row_if_empty_fields: tuple[str, ...] = ()
    skip_empty_cells: bool = True
    canonical_value_from_value: bool = False
    generic: bool = False


PROFILES = {
    "runtime_cards": ExportProfile(
        name="runtime_cards",
        sheet_name="7. EXPORT_RUNTIME",
        output_filename="EXPORT_RUNTIME.jsonl",
        output_format="jsonl",
        required_fields=("Card_ID", "Kártya név", "Típus", "Birodalom", "Magnitudó", "Aura", "Set_ID", "Collector_Number"),
        number_fields=("Magnitudó", "Aura", "ATK", "HP"),
    ),
    "decklists": ExportProfile(
        name="decklists",
        sheet_name="15. PRODUCT_DECKLISTS",
        output_filename="PRODUCT_DECKLISTS.jsonl",
        output_format="jsonl",
        required_fields=("Product_ID", "Deck_ID", "Card_ID", "Darabszám"),
        number_fields=("Darabszám",),
    ),
    "lookups_runtime": ExportProfile(
        name="lookups_runtime",
        sheet_name="5A. LOOKUPS_RUNTIME",
        output_filename="LOOKUPS_RUNTIME.jsonl",
        output_format="jsonl",
        required_fields=("Lookup_Group", "Value", "Status", "Used_For", "Sort_Order"),
        number_fields=("Sort_Order",),
        output_fields=LOOKUPS_FIELDS,
        skip_row_if_empty_fields=("Value",),
        canonical_value_from_value=True,
    ),
    "lookups_print_product": ExportProfile(
        name="lookups_print_product",
        sheet_name="5B. LOOKUPS_PRINT_PRODUCT",
        output_filename="LOOKUPS_PRINT_PRODUCT.jsonl",
        output_format="jsonl",
        required_fields=("Lookup_Group", "Value", "Status", "Used_For", "Sort_Order"),
        number_fields=("Sort_Order",),
        output_fields=LOOKUPS_FIELDS,
        skip_row_if_empty_fields=("Value",),
        canonical_value_from_value=True,
    ),
    "lookups_workflow_audit": ExportProfile(
        name="lookups_workflow_audit",
        sheet_name="5C. LOOKUPS_WORKFLOW_AUDIT",
        output_filename="LOOKUPS_WORKFLOW_AUDIT.jsonl",
        output_format="jsonl",
        required_fields=("Lookup_Group", "Value", "Status", "Used_For", "Sort_Order"),
        number_fields=("Sort_Order",),
        output_fields=LOOKUPS_FIELDS,
        skip_row_if_empty_fields=("Value",),
        canonical_value_from_value=True,
    ),
    "lookups_design_catalog": ExportProfile(
        name="lookups_design_catalog",
        sheet_name="5D. LOOKUPS_DESIGN_CATALOG",
        output_filename="LOOKUPS_DESIGN_CATALOG.jsonl",
        output_format="jsonl",
        required_fields=("Lookup_Group", "Value", "Status", "Used_For", "Sort_Order"),
        number_fields=("Sort_Order",),
        output_fields=LOOKUPS_FIELDS,
        skip_row_if_empty_fields=("Value",),
        canonical_value_from_value=True,
    ),
    "generic_sheet": ExportProfile(
        name="generic_sheet",
        sheet_name=None,
        output_filename=None,
        output_format="jsonl",
        generic=True,
    ),
}


def is_empty_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def normalize_value(value: Any, force_number: bool = False) -> Any:
    if value is None:
        return "none"
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if force_number and isinstance(value, str):
        cleaned = value.strip().replace(",", ".")
        if cleaned:
            try:
                number = float(cleaned)
            except ValueError:
                return value
            return int(number) if number.is_integer() else number
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return value


def load_headers(first_row: Sequence[Any] | None) -> list[str]:
    if first_row is None:
        raise ExportError("A kiválasztott munkalap üres, nincs fejlécsora.")

    headers = [str(normalize_value(value)) for value in first_row]
    if len(set(headers)) != len(headers):
        raise ExportError("A fejléc oszlopnevei nem lehetnek azonosak.")
    return headers


def iter_records(
    rows: Iterable[Sequence[Any]],
    headers: Sequence[str],
    profile: ExportProfile | None = None,
    warnings: list[str] | None = None,
):
    profile = profile or PROFILES["generic_sheet"]
    warnings = warnings if warnings is not None else []
    number_fields = set(profile.number_fields)

    for row_number, row in enumerate(rows, start=2):
        raw_by_header = {header: row[index] if index < len(row) else None for index, header in enumerate(headers)}
        skipped = False
        for field in profile.skip_row_if_empty_fields:
            if is_empty_value(raw_by_header.get(field)):
                warnings.append(f"{profile.name}: {row_number}. sor kihagyva: üres kötelező mező: {field}")
                skipped = True
                break
        if skipped:
            continue

        record = {}
        export_headers = profile.output_fields or tuple(headers)
        for header in export_headers:
            raw_value = raw_by_header.get(header)
            if profile.canonical_value_from_value and header == "Canonical_Value" and is_empty_value(raw_value):
                raw_value = raw_by_header.get("Value")
            if profile.skip_empty_cells and is_empty_value(raw_value):
                continue
            record[header] = normalize_value(raw_value, force_number=header in number_fields)

        for field in profile.required_fields:
            if field not in record or is_empty_value(record[field]):
                warnings.append(f"{profile.name}: {row_number}. sor: hiányzó kötelező mező: {field}")

        for field in number_fields:
            if field in record and record[field] != "none" and not isinstance(record[field], (int, float)):
                warnings.append(f"{profile.name}: {row_number}. sor: nem számként értelmezhető mező: {field}={record[field]!r}")

        if record:
            yield record


def safe_filename_part(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip().rstrip(".")
    return cleaned or "névtelen"


def default_output_path(xlsx_path: Path, sheet_name: str, output_format: str) -> Path:
    source_name = safe_filename_part(xlsx_path.stem)
    sheet_part = safe_filename_part(sheet_name)
    return DEFAULT_OUTPUT_DIR / f"{source_name}__{sheet_part}.{output_format}"


def profile_output_path(profile: ExportProfile) -> Path:
    if profile.output_filename is None:
        raise ExportError("A generic_sheet profilhoz kézzel kell kimeneti fájlt megadni.")
    return DEFAULT_OUTPUT_DIR / profile.output_filename


def resolve_export_options(
    profile_name: str | None,
    sheet_name: str | None,
    output_path: Path | None,
    output_format: str | None,
) -> tuple[ExportProfile, str, Path, str]:
    profile = PROFILES[profile_name or "generic_sheet"]

    if profile.generic:
        missing = [
            option
            for option, value in (("--sheet", sheet_name), ("--format", output_format), ("--output", output_path))
            if value is None
        ]
        if missing:
            raise ExportError(f"A generic_sheet profilhoz kötelező argumentum(ok): {', '.join(missing)}")
        return profile, sheet_name, output_path, output_format

    if sheet_name is not None and sheet_name != profile.sheet_name:
        raise ExportError(f"A(z) {profile.name} profil lapja kötött: {profile.sheet_name}")

    resolved_sheet = profile.sheet_name
    resolved_format = output_format or profile.output_format
    if resolved_format != profile.output_format:
        raise ExportError(f"A(z) {profile.name} profil kimeneti formátuma kötött: {profile.output_format}")
    resolved_output = output_path or profile_output_path(profile)
    return profile, resolved_sheet, resolved_output, resolved_format


def validate_profile_headers(headers: Sequence[str], profile: ExportProfile, warnings: list[str]) -> None:
    header_set = set(headers)
    for field in profile.required_fields:
        if field not in header_set:
            raise ExportError(f"A(z) {profile.name} profilhoz hiányzik a kötelező oszlop: {field}")
    for field in profile.number_fields:
        if field not in header_set:
            warnings.append(f"{profile.name}: hiányzó számmező oszlop: {field}")


def export_worksheet(
    xlsx_path: Path,
    sheet_name: str,
    output_path: Path,
    output_format: str,
    profile: ExportProfile | None = None,
) -> tuple[int, list[str]]:
    xlsx_path = Path(xlsx_path)
    output_path = Path(output_path)
    profile = profile or PROFILES["generic_sheet"]
    warnings: list[str] = []
    if xlsx_path.resolve() == output_path.resolve():
        raise ExportError("A kimeneti fájl nem lehet azonos a forrás .xlsx fájllal.")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ExportError(f"Nincs ilyen munkalap: {sheet_name}")

        rows = workbook[sheet_name].iter_rows(values_only=True)
        headers = load_headers(next(rows, None))
        validate_profile_headers(headers, profile, warnings)
        records = iter_records(rows, headers, profile=profile, warnings=warnings)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8", newline="") as output_file:
            if output_format == "jsonl":
                count = write_jsonl(output_file, records)
            else:
                count = write_tsv(output_file, headers, records)
        return count, warnings
    finally:
        workbook.close()


def write_warnings(output_path: Path, warnings: Sequence[str]) -> Path | None:
    warnings_path = output_path.with_suffix(output_path.suffix + ".warnings.txt")
    if not warnings:
        try:
            if warnings_path.exists():
                warnings_path.chmod(0o666)
                warnings_path.unlink()
        except OSError:
            pass
        return None
    if warnings_path.exists():
        warnings_path.chmod(0o666)
    warnings_path.write_text("\n".join(warnings) + "\n", encoding="utf-8")
    return warnings_path


def write_jsonl(output_file, records: Iterable[dict[str, Any]]) -> int:
    count = 0
    for record in records:
        output_file.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
        count += 1
    return count


def write_tsv(output_file, headers: Sequence[str], records: Iterable[dict[str, Any]]) -> int:
    writer = csv.DictWriter(output_file, fieldnames=headers, delimiter="\t", lineterminator="\n")
    writer.writeheader()
    count = 0
    for record in records:
        writer.writerow(record)
        count += 1
    return count


def list_sheets(xlsx_path: Path) -> list[str]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def find_xlsx_files() -> list[Path]:
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    files = {path.resolve() for path in SOURCE_DIR.glob("*.xlsx") if not path.name.startswith("~$")}
    return sorted(files, key=lambda path: path.name.casefold())


def validate_source_path(xlsx_path: Path) -> Path:
    resolved_path = xlsx_path.resolve()
    if resolved_path.parent != SOURCE_DIR.resolve():
        raise ExportError(f"A forrás XLSX fájlnak ebben a mappában kell lennie: {SOURCE_DIR}")
    return resolved_path


def choose_from_menu(title: str, options: Sequence[Any], label_func=str, input_func=input, print_func=print):
    if not options:
        raise ExportError(f"Nincs választható elem: {title}")

    print_func(f"\n{title}")
    for index, option in enumerate(options, start=1):
        print_func(f"  {index}. {label_func(option)}")
    print_func("  0. Kilépés")

    while True:
        answer = input_func("Választás: ").strip()
        if answer == "0":
            return None
        try:
            selected_index = int(answer) - 1
        except ValueError:
            selected_index = -1
        if 0 <= selected_index < len(options):
            return options[selected_index]
        print_func("Érvénytelen választás. Adj meg egy sorszámot.")


def interactive_menu(input_func=input, print_func=print) -> int:
    print_func("XLSX EXPORT")
    print_func(f"Forrásmappa: {SOURCE_DIR}")
    print_func("A kimeneti fájlok az exports mappába kerülnek.")

    xlsx_path = choose_from_menu(
        "Válaszd ki a forrás XLSX fájlt:",
        find_xlsx_files(),
        label_func=lambda path: path.name,
        input_func=input_func,
        print_func=print_func,
    )
    if xlsx_path is None:
        return 0

    profile_name = choose_from_menu(
        "Válaszd ki az exportprofilt:",
        tuple(PROFILES),
        input_func=input_func,
        print_func=print_func,
    )
    if profile_name is None:
        return 0

    sheet_name = None
    output_format = None
    output_path = None
    if profile_name == "generic_sheet":
        sheet_name = choose_from_menu(
            "Válaszd ki az exportálandó munkalapot:",
            list_sheets(xlsx_path),
            input_func=input_func,
            print_func=print_func,
        )
        if sheet_name is None:
            return 0

        output_format = choose_from_menu(
            "Válaszd ki a kimeneti formátumot:",
            ("jsonl", "tsv"),
            input_func=input_func,
            print_func=print_func,
        )
        if output_format is None:
            return 0
        output_path = default_output_path(xlsx_path, sheet_name, output_format)

    profile, sheet_name, output_path, output_format = resolve_export_options(profile_name, sheet_name, output_path, output_format)
    count, warnings = export_worksheet(xlsx_path, sheet_name, output_path, output_format, profile=profile)
    warnings_path = write_warnings(output_path, warnings)
    print_func(f"\nExportált sorok száma: {count}")
    print_func(f"Kimeneti fájl: {output_path.resolve()}")
    print_warning_summary(warnings, warnings_path, print_func=print_func)
    return 0


def print_warning_summary(warnings: Sequence[str], warnings_path: Path | None, print_func=print) -> None:
    if not warnings:
        print_func("Warningok: 0")
        return
    print_func(f"Warningok: {len(warnings)}")
    for warning in warnings[:10]:
        print_func(f"  - {warning}")
    if len(warnings) > 10:
        print_func(f"  ... további warningok: {len(warnings) - 10}")
    if warnings_path is not None:
        print_func(f"Warning fájl: {warnings_path.resolve()}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="XLSX munkalap exportálása JSONL vagy TSV formátumba.")
    parser.add_argument("xlsx", nargs="?", type=Path, help="A forrás .xlsx fájl")
    parser.add_argument("--profile", choices=tuple(PROFILES), default="generic_sheet", help="Exportprofil")
    parser.add_argument("--list-sheets", action="store_true", help="A munkalapok kilistázása")
    parser.add_argument("--sheet", help="Az exportálandó munkalap neve")
    parser.add_argument("--format", choices=("jsonl", "tsv"), dest="output_format", help="Kimeneti formátum")
    parser.add_argument(
        "--output",
        type=Path,
        help="Egyedi kimeneti fájl. Ha nincs megadva, az exports mappába készül automatikus névvel.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        if args.xlsx is None:
            return interactive_menu()
        args.xlsx = validate_source_path(args.xlsx)
        if not args.xlsx.is_file():
            raise ExportError(f"A forrásfájl nem található: {args.xlsx}")
        if args.xlsx.suffix.lower() != ".xlsx":
            raise ExportError("A forrásfájlnak .xlsx kiterjesztésűnek kell lennie.")

        if args.list_sheets:
            for sheet_name in list_sheets(args.xlsx):
                print(sheet_name)
            return 0

        profile, sheet_name, output_path, output_format = resolve_export_options(
            args.profile,
            args.sheet,
            args.output,
            args.output_format,
        )
        count, warnings = export_worksheet(args.xlsx, sheet_name, output_path, output_format, profile=profile)
        warnings_path = write_warnings(output_path, warnings)
        print(f"Exportált sorok száma: {count}")
        print(f"Kimeneti fájl: {output_path.resolve()}")
        print_warning_summary(warnings, warnings_path)
        return 0
    except (ExportError, OSError, ValueError) as exc:
        print(f"Hiba: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"Hiba az XLSX fájl feldolgozásakor: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
