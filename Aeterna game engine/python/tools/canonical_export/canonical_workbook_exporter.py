"""Manifest-driven canonical JSON exporter for AETERNA workbooks."""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import uuid
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook


CANONICAL_FORMAT = "aeterna_canonical_table_v1"
SUPPORTED_EXPORT_FORMATS = {"json"}
SENTINEL_NULL = "#NULL"
SENTINEL_TBD = "#TBD"


@dataclass(frozen=True)
class Diagnostic:
    code: str
    message: str
    workbook: str
    sheet: str | None = None
    row: int | None = None
    field: str | None = None
    rule: str | None = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "code": self.code,
            "message": self.message,
            "workbook": self.workbook,
            "sheet": self.sheet,
            "row": self.row,
            "field": self.field,
            "rule": self.rule,
        }


class CanonicalExportError(Exception):
    def __init__(self, diagnostics: Sequence[Diagnostic]):
        super().__init__(f"Canonical export blocked by {len(diagnostics)} diagnostic(s).")
        self.diagnostics = tuple(diagnostics)


@dataclass(frozen=True)
class FieldSchema:
    field_id: str
    table_id: str
    field_name: str
    column_order: int
    required_mode: str
    nullable: bool
    null_handling: str


@dataclass(frozen=True)
class TableSchema:
    table_id: str
    sheet_name: str
    primary_key: str


@dataclass(frozen=True)
class ManifestEntry:
    table_id: str
    export_file: str
    export_format: str
    export_order: int


@dataclass
class PreparedPackage:
    workbook_path: Path
    package_id: str
    package_directory_name: str
    null_sentinel: str
    tbd_sentinel: str
    entries: list[ManifestEntry]
    records: dict[str, list[dict[str, Any]]]


def _diagnostic(
    diagnostics: list[Diagnostic],
    workbook: Path,
    code: str,
    message: str,
    *,
    sheet: str | None = None,
    row: int | None = None,
    field: str | None = None,
    rule: str | None = None,
) -> None:
    diagnostics.append(Diagnostic(code, message, workbook.name, sheet, row, field, rule))


def _rows(worksheet) -> tuple[list[str], list[tuple[int, tuple[Any, ...]]]]:
    values = worksheet.iter_rows(values_only=True)
    try:
        first = next(values)
    except StopIteration:
        return [], []
    headers = ["" if value is None else str(value).strip() for value in first]
    last = max((index for index, value in enumerate(headers) if value), default=-1)
    headers = headers[: last + 1]
    data: list[tuple[int, tuple[Any, ...]]] = []
    for row_number, row in enumerate(values, start=2):
        row = tuple(row[: len(headers)])
        if any(value is not None and (not isinstance(value, str) or value.strip()) for value in row):
            data.append((row_number, row))
    return headers, data


def _raw_records(worksheet) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    headers, rows = _rows(worksheet)
    return headers, [(number, dict(zip(headers, row))) for number, row in rows]


def _is_true(value: Any) -> bool:
    return value is True or (isinstance(value, str) and value.strip().casefold() == "true")


def _integer(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def _json_value(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return value


def _safe_export_file(value: Any) -> str | None:
    if (
        not isinstance(value, str)
        or not value
        or value != value.strip()
        or value in {".", ".."}
        or "/" in value
        or "\\" in value
        or re.match(r"^[A-Za-z]:", value) is not None
        or Path(value).is_absolute()
        or value != Path(value).name
    ):
        return None
    if Path(value).suffix != ".json":
        return None
    return value


def _safe_package_directory(value: Any) -> str | None:
    if (
        not isinstance(value, str)
        or not value
        or value != value.strip()
        or value in {".", ".."}
        or "/" in value
        or "\\" in value
        or re.match(r"^[A-Za-z]:", value) is not None
        or Path(value).is_absolute()
        or value != Path(value).name
    ):
        return None
    return value


def _write_json(path: Path, value: Any) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, separators=(",", ": ")) + "\n",
        encoding="utf-8",
        newline="\n",
    )


def _meta_values(workbook, workbook_path: Path, diagnostics: list[Diagnostic]) -> dict[str, Any]:
    if "META" not in workbook.sheetnames:
        _diagnostic(diagnostics, workbook_path, "CANONICAL_META_SHEET_MISSING", "META worksheet is missing.")
        return {}
    headers, rows = _raw_records(workbook["META"])
    if headers[:2] != ["key", "value"]:
        _diagnostic(
            diagnostics,
            workbook_path,
            "CANONICAL_META_HEADER_INVALID",
            "META must begin with key and value columns.",
            sheet="META",
            row=1,
        )
        return {}
    result: dict[str, Any] = {}
    for row_number, record in rows:
        key = record.get("key")
        if not isinstance(key, str) or not key.strip():
            _diagnostic(diagnostics, workbook_path, "CANONICAL_META_KEY_INVALID", "META key is empty.", sheet="META", row=row_number, field="key")
        elif key in result:
            _diagnostic(diagnostics, workbook_path, "CANONICAL_META_KEY_DUPLICATE", "META key is duplicated.", sheet="META", row=row_number, field="key")
        else:
            result[key] = record.get("value")
    return result


def _prepare_workbook(workbook_path: Path, production: bool) -> tuple[PreparedPackage | None, list[Diagnostic]]:
    diagnostics: list[Diagnostic] = []
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    except Exception as exception:
        _diagnostic(diagnostics, workbook_path, "CANONICAL_WORKBOOK_INVALID", f"Workbook could not be read: {exception}")
        return None, diagnostics

    try:
        required_contract_sheets = {"META", "EXPORT_MANIFEST", "SCHEMA_TABLES", "SCHEMA_FIELDS", "VALIDATION_RULES"}
        for sheet_name in sorted(required_contract_sheets):
            if sheet_name not in workbook.sheetnames:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_CONTRACT_SHEET_MISSING", "Required contract worksheet is missing.", sheet=sheet_name)
        if diagnostics:
            return None, diagnostics

        meta = _meta_values(workbook, workbook_path, diagnostics)
        package_id = meta.get("workbook_id")
        directory_name = meta.get("export_directory_name")
        declared_manifest = meta.get("export_manifest_file")
        null_sentinel = meta.get("null_sentinel")
        tbd_sentinel = meta.get("tbd_sentinel")
        if not isinstance(package_id, str) or not package_id.strip():
            _diagnostic(diagnostics, workbook_path, "CANONICAL_PACKAGE_ID_INVALID", "META.workbook_id must be a non-empty string.", sheet="META", field="workbook_id")
        if _safe_package_directory(directory_name) is None:
            _diagnostic(diagnostics, workbook_path, "CANONICAL_PACKAGE_DIRECTORY_INVALID", "META.export_directory_name must be a safe directory name.", sheet="META", field="export_directory_name")
        if not isinstance(null_sentinel, str) or not null_sentinel or null_sentinel != SENTINEL_NULL:
            _diagnostic(diagnostics, workbook_path, "CANONICAL_NULL_SENTINEL_INVALID", "META.null_sentinel must be the canonical v1 literal #NULL.", sheet="META", field="null_sentinel")
        if not isinstance(tbd_sentinel, str) or not tbd_sentinel or tbd_sentinel != SENTINEL_TBD:
            _diagnostic(diagnostics, workbook_path, "CANONICAL_TBD_SENTINEL_INVALID", "META.tbd_sentinel must be the canonical v1 literal #TBD.", sheet="META", field="tbd_sentinel")

        table_headers, table_rows = _raw_records(workbook["SCHEMA_TABLES"])
        required_table_headers = ["table_id", "sheet_name", "primary_key", "status"]
        if any(header not in table_headers for header in required_table_headers):
            _diagnostic(diagnostics, workbook_path, "CANONICAL_SCHEMA_TABLES_HEADER_INVALID", "SCHEMA_TABLES is missing required columns.", sheet="SCHEMA_TABLES", row=1)
        table_schemas: dict[str, TableSchema] = {}
        for row_number, record in table_rows:
            if record.get("status") != "active":
                continue
            table_id = record.get("table_id")
            sheet_name = record.get("sheet_name")
            primary_key = record.get("primary_key")
            if not all(isinstance(value, str) and value for value in (table_id, sheet_name, primary_key)):
                _diagnostic(diagnostics, workbook_path, "CANONICAL_SCHEMA_TABLE_INVALID", "Active table schema is incomplete.", sheet="SCHEMA_TABLES", row=row_number)
                continue
            if table_id in table_schemas:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_TABLE_ID_DUPLICATE", "Active table_id is duplicated.", sheet="SCHEMA_TABLES", row=row_number, field="table_id")
                continue
            table_schemas[table_id] = TableSchema(table_id, sheet_name, primary_key)

        field_headers, field_rows = _raw_records(workbook["SCHEMA_FIELDS"])
        required_field_headers = ["field_id", "table_id", "field_name", "column_order", "required_mode", "nullable", "null_handling", "status"]
        if any(header not in field_headers for header in required_field_headers):
            _diagnostic(diagnostics, workbook_path, "CANONICAL_SCHEMA_FIELDS_HEADER_INVALID", "SCHEMA_FIELDS is missing required columns.", sheet="SCHEMA_FIELDS", row=1)
        seen_field_ids: set[str] = set()
        seen_table_fields: set[tuple[str, str]] = set()
        fields_by_table: dict[str, list[FieldSchema]] = {}
        for row_number, record in field_rows:
            field_id = record.get("field_id")
            table_id = record.get("table_id")
            field_name = record.get("field_name")
            if not isinstance(field_id, str) or not field_id:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_FIELD_ID_INVALID", "field_id is empty.", sheet="SCHEMA_FIELDS", row=row_number, field="field_id")
                continue
            if field_id in seen_field_ids:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_FIELD_ID_DUPLICATE", "field_id is duplicated.", sheet="SCHEMA_FIELDS", row=row_number, field="field_id")
            seen_field_ids.add(field_id)
            pair = (str(table_id), str(field_name))
            if pair in seen_table_fields:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_TABLE_FIELD_DUPLICATE", "table_id + field_name is duplicated.", sheet="SCHEMA_FIELDS", row=row_number, field="field_name")
            seen_table_fields.add(pair)
            if record.get("status") != "active":
                continue
            order = _integer(record.get("column_order"))
            if not isinstance(table_id, str) or not isinstance(field_name, str) or order is None:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_ACTIVE_FIELD_INVALID", "Active field schema is incomplete.", sheet="SCHEMA_FIELDS", row=row_number)
                continue
            fields_by_table.setdefault(table_id, []).append(
                FieldSchema(field_id, table_id, field_name, order, str(record.get("required_mode")), _is_true(record.get("nullable")), str(record.get("null_handling")))
            )
        for table_id, fields in fields_by_table.items():
            fields.sort(key=lambda item: item.column_order)
            orders = [item.column_order for item in fields]
            if orders != list(range(1, len(fields) + 1)):
                _diagnostic(diagnostics, workbook_path, "CANONICAL_COLUMN_ORDER_INVALID", "Active field column_order must be unique and gapless from one.", sheet="SCHEMA_FIELDS", field=table_id)

        manifest_headers, manifest_rows = _raw_records(workbook["EXPORT_MANIFEST"])
        required_manifest_headers = ["table_id", "export_enabled", "export_file", "export_format", "export_order"]
        if any(header not in manifest_headers for header in required_manifest_headers):
            _diagnostic(diagnostics, workbook_path, "CANONICAL_MANIFEST_HEADER_INVALID", "EXPORT_MANIFEST is missing required columns.", sheet="EXPORT_MANIFEST", row=1)
        entries: list[ManifestEntry] = []
        files: set[str] = set()
        orders: set[int] = set()
        for row_number, record in manifest_rows:
            if not _is_true(record.get("export_enabled")):
                continue
            table_id = record.get("table_id")
            export_file = _safe_export_file(record.get("export_file"))
            export_format = record.get("export_format")
            export_order = _integer(record.get("export_order"))
            if not isinstance(table_id, str) or table_id not in table_schemas:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_MANIFEST_TABLE_UNKNOWN", "Enabled manifest table_id has no active schema.", sheet="EXPORT_MANIFEST", row=row_number, field="table_id")
                continue
            if export_file is None:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_EXPORT_FILE_INVALID", "Enabled export_file is not a safe JSON filename.", sheet="EXPORT_MANIFEST", row=row_number, field="export_file")
                continue
            if export_format not in SUPPORTED_EXPORT_FORMATS:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_EXPORT_FORMAT_UNSUPPORTED", "Enabled export_format is unsupported.", sheet="EXPORT_MANIFEST", row=row_number, field="export_format")
                continue
            if export_order is None or export_order < 1:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_EXPORT_ORDER_INVALID", "Enabled export_order must be a positive integer.", sheet="EXPORT_MANIFEST", row=row_number, field="export_order")
                continue
            if export_file in files:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_EXPORT_FILE_DUPLICATE", "Enabled export_file is duplicated.", sheet="EXPORT_MANIFEST", row=row_number, field="export_file")
            if export_order in orders:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_EXPORT_ORDER_DUPLICATE", "Enabled export_order is duplicated.", sheet="EXPORT_MANIFEST", row=row_number, field="export_order")
            files.add(export_file)
            orders.add(export_order)
            entries.append(ManifestEntry(table_id, export_file, export_format, export_order))
        entries.sort(key=lambda item: item.export_order)
        manifest_entries = [entry for entry in entries if entry.table_id == "export_manifest"]
        if len(manifest_entries) != 1 or manifest_entries[0].export_order != min((entry.export_order for entry in entries), default=0):
            _diagnostic(diagnostics, workbook_path, "CANONICAL_MANIFEST_ENTRY_INVALID", "export_manifest must be enabled exactly once and load first.", sheet="EXPORT_MANIFEST")
        if manifest_entries and declared_manifest != manifest_entries[0].export_file:
            _diagnostic(diagnostics, workbook_path, "CANONICAL_MANIFEST_FILENAME_MISMATCH", "META.export_manifest_file contradicts EXPORT_MANIFEST.", sheet="META", field="export_manifest_file")

        exported_records: dict[str, list[dict[str, Any]]] = {}
        for entry in entries:
            schema = table_schemas[entry.table_id]
            fields = fields_by_table.get(entry.table_id, [])
            if schema.sheet_name not in workbook.sheetnames:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_EXPORTED_SHEET_MISSING", "Manifest-referenced worksheet is missing.", sheet=schema.sheet_name, field=entry.table_id)
                continue
            expected_headers = [field.field_name for field in fields]
            actual_headers, source_rows = _rows(workbook[schema.sheet_name])
            if actual_headers != expected_headers:
                _diagnostic(diagnostics, workbook_path, "CANONICAL_HEADER_MISMATCH", "Worksheet header does not match its active schema.", sheet=schema.sheet_name, row=1, rule="headers_match_active_schema")
                continue
            records: list[dict[str, Any]] = []
            primary_values: set[Any] = set()
            for row_number, values in source_rows:
                record: dict[str, Any] = {}
                raw_record = dict(zip(expected_headers, values))
                for field_schema, raw in zip(fields, values):
                    value = raw
                    sentinel_declaration = (
                        entry.table_id == "meta"
                        and field_schema.field_name == "value"
                        and raw_record.get("key") in {"null_sentinel", "tbd_sentinel"}
                    )
                    if value == tbd_sentinel and production and not sentinel_declaration:
                        _diagnostic(diagnostics, workbook_path, "CANONICAL_TBD_FORBIDDEN", "#TBD is forbidden in production export.", sheet=schema.sheet_name, row=row_number, field=field_schema.field_name, rule="production_tbd_policy")
                    if value == null_sentinel and not sentinel_declaration:
                        if field_schema.nullable:
                            value = None
                        else:
                            _diagnostic(diagnostics, workbook_path, "CANONICAL_NULL_FORBIDDEN", "#NULL is forbidden by the active field schema.", sheet=schema.sheet_name, row=row_number, field=field_schema.field_name, rule="nullable")
                    elif value is None or (isinstance(value, str) and not value.strip()):
                        if field_schema.nullable and field_schema.null_handling == "blank_is_null":
                            value = None
                        else:
                            _diagnostic(diagnostics, workbook_path, "CANONICAL_REQUIRED_VALUE_MISSING", "Blank value is not allowed by the active field schema.", sheet=schema.sheet_name, row=row_number, field=field_schema.field_name, rule=field_schema.required_mode)
                    record[field_schema.field_name] = _json_value(value)
                primary_value = record.get(schema.primary_key)
                if primary_value is None or primary_value == "":
                    _diagnostic(diagnostics, workbook_path, "CANONICAL_PRIMARY_KEY_MISSING", "Primary record ID is missing.", sheet=schema.sheet_name, row=row_number, field=schema.primary_key)
                elif primary_value in primary_values:
                    _diagnostic(diagnostics, workbook_path, "CANONICAL_PRIMARY_KEY_DUPLICATE", "Primary record ID is duplicated.", sheet=schema.sheet_name, row=row_number, field=schema.primary_key)
                else:
                    primary_values.add(primary_value)
                records.append(record)
            exported_records[entry.table_id] = records

        _validate_safe_uniqueness_rules(workbook, workbook_path, exported_records, fields_by_table, diagnostics, null_sentinel)
        if diagnostics:
            return None, diagnostics
        assert isinstance(package_id, str) and isinstance(directory_name, str)
        assert isinstance(null_sentinel, str) and isinstance(tbd_sentinel, str)
        return PreparedPackage(workbook_path, package_id, directory_name, null_sentinel, tbd_sentinel, entries, exported_records), diagnostics
    finally:
        workbook.close()


def _validate_safe_uniqueness_rules(
    workbook,
    workbook_path: Path,
    records_by_table: dict[str, list[dict[str, Any]]],
    fields_by_table: dict[str, list[FieldSchema]],
    diagnostics: list[Diagnostic],
    null_sentinel: Any,
) -> None:
    _, rules = _raw_records(workbook["VALIDATION_RULES"])
    field_names = {field.field_id: field.field_name for fields in fields_by_table.values() for field in fields}
    unique_by = re.compile(r'^unique_by\(\[([^]]+)\]\)$')
    for row_number, rule in rules:
        if rule.get("status") != "active" or not _is_true(rule.get("blocking")) or rule.get("validation_kind_id") != "uniqueness":
            continue
        table_id = rule.get("target_table_id")
        records = records_by_table.get(str(table_id))
        if records is None:
            continue
        expression = rule.get("condition_expression")
        names: list[str] = []
        if expression == null_sentinel:
            name = field_names.get(str(rule.get("target_field_id")))
            if name:
                names = [name]
        elif isinstance(expression, str):
            match = unique_by.fullmatch(expression.strip())
            if match:
                names = [part.strip().strip('"') for part in match.group(1).split(",")]
        if not names:
            continue
        seen: set[tuple[Any, ...]] = set()
        for index, record in enumerate(records, start=2):
            key = tuple(record.get(name) for name in names)
            if key in seen:
                _diagnostic(diagnostics, workbook_path, str(rule.get("error_code") or "CANONICAL_UNIQUENESS_RULE_FAILED"), str(rule.get("message") or "Active uniqueness rule failed."), sheet=str(table_id), row=index, field=",".join(names), rule=str(rule.get("validation_rule_id")))
            seen.add(key)


def export_canonical_workbooks(
    workbook_paths: Sequence[Path | str],
    output_root: Path | str,
    *,
    production: bool = True,
    diagnostics_path: Path | str | None = None,
) -> tuple[Path, ...]:
    paths = [Path(path).resolve() for path in workbook_paths]
    prepared: list[PreparedPackage] = []
    diagnostics: list[Diagnostic] = []
    for path in paths:
        package, workbook_diagnostics = _prepare_workbook(path, production)
        diagnostics.extend(workbook_diagnostics)
        if package is not None:
            prepared.append(package)
    directory_names = [package.package_directory_name for package in prepared]
    if len(directory_names) != len(set(directory_names)):
        diagnostics.append(Diagnostic("CANONICAL_PACKAGE_DIRECTORY_DUPLICATE", "Multiple workbooks declare the same package directory.", "<multiple>"))
    if diagnostics:
        if diagnostics_path is not None:
            report_path = Path(diagnostics_path)
            report_path.parent.mkdir(parents=True, exist_ok=True)
            _write_json(report_path, {"status": "blocked", "diagnostics": [item.as_dict() for item in diagnostics]})
        raise CanonicalExportError(diagnostics)

    root = Path(output_root).resolve()
    root.mkdir(parents=True, exist_ok=True)
    staged: list[tuple[PreparedPackage, Path]] = []
    backups: list[tuple[Path, Path]] = []
    published: list[Path] = []
    preserve_temp_root = False
    temp_root = root / (".canonical-export-" + uuid.uuid4().hex)
    temp_root.mkdir()
    try:
        # Stage every package before changing any published target.
        for package in prepared:
            package_root = temp_root / package.package_directory_name
            package_root.mkdir()
            for entry in package.entries:
                envelope = {
                    "canonical_format": CANONICAL_FORMAT,
                    "package_id": package.package_id,
                    "table_id": entry.table_id,
                    "records": package.records[entry.table_id],
                }
                _write_json(package_root / entry.export_file, envelope)
            staged.append((package, package_root))

        # Back up every existing target before publishing any staged package.
        for package, staged_root in staged:
            target = root / package.package_directory_name
            backup = temp_root / (package.package_directory_name + ".previous")
            if target.exists():
                os.replace(target, backup)
                backups.append((target, backup))

        # Publish every package as one transaction. Backups remain intact until
        # every staged package has reached its final target.
        for package, staged_root in staged:
            target = root / package.package_directory_name
            os.replace(staged_root, target)
            published.append(target)
    except Exception as publish_error:
        # Remove every newly published package, including targets that did not
        # exist before this transaction, then restore all available backups.
        rollback_errors: list[Exception] = []
        for target in reversed(published):
            try:
                if target.exists():
                    shutil.rmtree(target)
            except Exception as remove_error:
                rollback_errors.append(remove_error)
        for target, backup in reversed(backups):
            try:
                if target.exists():
                    shutil.rmtree(target)
                if backup.exists():
                    os.replace(backup, target)
            except Exception as restore_error:
                rollback_errors.append(restore_error)
                preserve_temp_root = preserve_temp_root or backup.exists()
        if rollback_errors:
            raise CanonicalExportError(
                [
                    Diagnostic(
                        "CANONICAL_PUBLISH_ROLLBACK_FAILED",
                        f"Canonical publish failed and rollback could not restore every package: {rollback_errors[0]}",
                        "<multiple>",
                    )
                ]
            ) from publish_error
        raise
    finally:
        if not preserve_temp_root:
            shutil.rmtree(temp_root, ignore_errors=True)
    if diagnostics_path is not None:
        report_path = Path(diagnostics_path)
        report_path.parent.mkdir(parents=True, exist_ok=True)
        _write_json(report_path, {"status": "success", "packages": [str(path) for path in published], "diagnostics": []})
    return tuple(published)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbooks", nargs="+", type=Path, help="REGISTRY.xlsx and/or CARDDATABASE.xlsx")
    parser.add_argument("--output-root", required=True, type=Path, help="Parent directory for declared package directories")
    parser.add_argument("--diagnostics", type=Path, help="Structured diagnostics JSON path")
    parser.add_argument("--development", action="store_true", help="Permit schema-authorized #TBD values (never use for production input)")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        packages = export_canonical_workbooks(args.workbooks, args.output_root, production=not args.development, diagnostics_path=args.diagnostics)
    except CanonicalExportError as exception:
        for item in exception.diagnostics:
            print(json.dumps(item.as_dict(), ensure_ascii=False), file=__import__("sys").stderr)
        return 2
    for package in packages:
        print(package)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
