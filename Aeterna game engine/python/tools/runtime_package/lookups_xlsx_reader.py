"""Read runtime lookup rows from the canonical LOOKUPS.xlsx workbook."""

from __future__ import annotations

import re
from typing import Any

from openpyxl import load_workbook


RUNTIME_LOOKUP_SHEETS = ("RUNTIME_CORE", "RUNTIME_ABILITY")
LOOKUP_HEADERS = (
    "Lookup_Group",
    "Value",
    "Label_HU",
    "Status",
    "Canonical_Value",
    "Used_For",
    "Sort_Order",
    "Source",
    "Notes",
)


class LookupsXlsxReaderError(Exception):
    """Raised when the canonical LOOKUPS.xlsx structure is not readable."""


def load_runtime_lookups_from_xlsx(xlsx_path, sheet_names=RUNTIME_LOOKUP_SHEETS, active_only=True):
    """Return builder-compatible lookup records from runtime lookup sheets.

    The returned lookup dictionaries match the shape consumed by the runtime
    package builder's lookup input path.
    """
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        lookups = []
        summary = {
            "sheets_read": [],
            "records_read": 0,
            "lookups_loaded": 0,
            "skipped_empty_value": 0,
            "skipped_inactive": 0,
            "diagnostics": [],
        }

        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                raise LookupsXlsxReaderError("Missing LOOKUPS.xlsx sheet: %s" % sheet_name)

            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            headers = _read_headers(header_row, sheet_name)
            summary["sheets_read"].append(sheet_name)

            for row_number, row in enumerate(rows, start=2):
                if not row or not any(_has_value(value) for value in row):
                    continue

                summary["records_read"] += 1
                source_row = _row_to_dict(headers, row)
                value = _text_value(source_row.get("Value"))
                if value == "none":
                    summary["skipped_empty_value"] += 1
                    continue

                status = _text_value(source_row.get("Status"), default="active")
                if active_only and status.casefold() != "active":
                    summary["skipped_inactive"] += 1
                    continue

                lookups.append(_to_runtime_lookup(source_row))

        summary["lookups_loaded"] = len(lookups)
        return {"lookups": lookups, "summary": summary}
    finally:
        workbook.close()


def _read_headers(header_row, sheet_name):
    if header_row is None:
        raise LookupsXlsxReaderError("Missing header row in sheet: %s" % sheet_name)
    headers = [_text_value(value, default="") for value in header_row]
    missing = [header for header in LOOKUP_HEADERS if header not in headers]
    if missing:
        raise LookupsXlsxReaderError(
            "Missing LOOKUPS.xlsx columns in %s: %s" % (sheet_name, ", ".join(missing))
        )
    return headers


def _row_to_dict(headers, row):
    return {
        header: row[index] if index < len(row) else None
        for index, header in enumerate(headers)
        if header
    }


def _to_runtime_lookup(row):
    value = _text_value(row.get("Value"))
    canonical_value = _text_value(row.get("Canonical_Value"), default=value)
    return {
        "lookup_group": _canonical_lookup_group(row.get("Lookup_Group")),
        "value": value,
        "label_hu": _text_value(row.get("Label_HU"), default=value),
        "status": _text_value(row.get("Status"), default="active"),
        "canonical_value": canonical_value,
        "used_for": _list_value(row.get("Used_For")),
        "sort_order": _number_value(row.get("Sort_Order")),
        "source": _text_value(row.get("Source")),
        "notes": _text_value(row.get("Notes")),
    }


def _canonical_lookup_group(value):
    text = _text_value(value)
    if text == "none":
        return text
    text = re.sub(r"[\s\-]+", "_", text.strip())
    return text.casefold()


def _list_value(value):
    text = _text_value(value)
    if text == "none":
        return []
    return [part.strip() for part in text.replace(",", ";").split(";") if part.strip()]


def _number_value(value):
    if value is None or value == "":
        return "none"
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return "none"
        try:
            number = float(text)
        except ValueError:
            return text
        return int(number) if number.is_integer() else number
    return value


def _text_value(value: Any, default="none"):
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def _has_value(value):
    return value is not None and str(value).strip() != ""
