"""Read legacy normalization aliases from LOOKUPS.xlsx."""

from __future__ import annotations

import re
from typing import Any

from openpyxl import load_workbook


LEGACY_ALIAS_SHEET = "RUNTIME_LEGACY_ALIAS"
LEGACY_ALIAS_HEADERS = (
    "Lookup_Group",
    "Value",
    "Label_HU",
    "Status",
    "Canonical_Value",
    "Used_For",
    "Sort_Order",
    "Source",
    "Notes",
)
AUDIT_REQUIRED = "audit_required"


class RuntimeLegacyAliasesReaderError(Exception):
    """Raised when the RUNTIME_LEGACY_ALIAS sheet is not readable."""


def load_runtime_legacy_aliases_from_xlsx(xlsx_path, sheet_name=LEGACY_ALIAS_SHEET):
    """Return normalized legacy alias records from LOOKUPS.xlsx."""
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeLegacyAliasesReaderError("Missing LOOKUPS.xlsx sheet: %s" % sheet_name)

        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = _read_headers(next(rows, None), sheet_name)

        aliases = []
        summary = {
            "sheet_read": sheet_name,
            "records_read": 0,
            "aliases_loaded": 0,
            "skipped_empty_value": 0,
            "requires_audit_count": 0,
            "normalization_allowed_count": 0,
            "diagnostics": [],
        }

        for row in rows:
            if not row or not any(_has_value(value) for value in row):
                continue

            summary["records_read"] += 1
            source_row = _row_to_dict(headers, row)
            alias_value = _text_value(source_row.get("Value"))
            if alias_value == "none":
                summary["skipped_empty_value"] += 1
                continue

            alias = _to_legacy_alias(source_row)
            aliases.append(alias)
            if alias["requires_audit"]:
                summary["requires_audit_count"] += 1
            if alias["normalization_allowed"]:
                summary["normalization_allowed_count"] += 1

        summary["aliases_loaded"] = len(aliases)
        return {"aliases": aliases, "summary": summary}
    finally:
        workbook.close()


def _read_headers(header_row, sheet_name):
    if header_row is None:
        raise RuntimeLegacyAliasesReaderError("Missing header row in sheet: %s" % sheet_name)
    headers = [_text_value(value, default="") for value in header_row]
    missing = [header for header in LEGACY_ALIAS_HEADERS if header not in headers]
    if missing:
        raise RuntimeLegacyAliasesReaderError(
            "Missing RUNTIME_LEGACY_ALIAS columns in %s: %s" % (sheet_name, ", ".join(missing))
        )
    return headers


def _row_to_dict(headers, row):
    return {
        header: row[index] if index < len(row) else None
        for index, header in enumerate(headers)
        if header
    }


def _to_legacy_alias(row):
    alias_value = _text_value(row.get("Value"))
    canonical_value = _text_value(row.get("Canonical_Value"))
    requires_audit = canonical_value == "none" or canonical_value.casefold() == AUDIT_REQUIRED
    return {
        "lookup_group": _canonical_lookup_group(row.get("Lookup_Group")),
        "alias_value": alias_value,
        "label_hu": _text_value(row.get("Label_HU"), default=alias_value),
        "canonical_value": canonical_value,
        "status": _text_value(row.get("Status"), default="legacy"),
        "used_for": _list_value(row.get("Used_For")),
        "sort_order": _number_value(row.get("Sort_Order")),
        "source": _text_value(row.get("Source")),
        "notes": _text_value(row.get("Notes")),
        "requires_audit": requires_audit,
        "normalization_allowed": not requires_audit,
    }


def _canonical_lookup_group(value):
    text = _text_value(value)
    if text == "none":
        return text
    text = re.sub(r"[\s\-]+", "_", text.strip())
    return text.casefold()


def _list_value(value):
    text = _text_value(value)
    if text == "none":
        return []
    return [part.strip() for part in text.replace(",", ";").split(";") if part.strip()]


def _number_value(value):
    if value is None or value == "":
        return "none"
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return "none"
        try:
            number = float(text)
        except ValueError:
            return text
        return int(number) if number.is_integer() else number
    return value


def _text_value(value: Any, default="none"):
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def _has_value(value):
    return value is not None and str(value).strip() != ""
