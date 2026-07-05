import importlib.util
import json
import shutil
import sys
import tempfile
import uuid
import unittest
from pathlib import Path

from openpyxl import Workbook


SCRIPT_PATH = (
    Path(__file__).resolve().parents[1]
    / "tools"
    / "runtime_package"
    / "smoke_real_export_runtime_package.py"
)
MAPPER_PATH = (
    Path(__file__).resolve().parents[1]
    / "tools"
    / "runtime_package"
    / "runtime_card_mapper.py"
)
BUILDER_PATH = (
    Path(__file__).resolve().parents[1]
    / "tools"
    / "runtime_package"
    / "build_sample_runtime_package.py"
)
ENGINE_PYTHON_DIR = Path(__file__).resolve().parents[1]


def _load_module(module_name, path):
    spec = importlib.util.spec_from_file_location(module_name, path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


class TestSmokeRealExportRuntimePackage(unittest.TestCase):
    def setUp(self):
        self.runner = _load_module("smoke_real_export_runtime_package", SCRIPT_PATH)
        self.xlsx_export = self.runner.XLSX_EXPORT
        self.mapper = _load_module("runtime_card_mapper", MAPPER_PATH)
        self.builder = _load_module("build_sample_runtime_package", BUILDER_PATH)
        self.temp_dir = Path(tempfile.gettempdir()) / ("real_export_runtime_package_smoke_%s" % uuid.uuid4().hex)
        self.temp_dir.mkdir(parents=True)

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)
        self.assertFalse(self.temp_dir.exists(), "Real export smoke temp cleanup left directory: %s" % self.temp_dir)

    def test_smoke_runner_builds_partial_package_from_xlsx(self):
        xlsx_path = self.temp_dir / "aeterna_cards.xlsx"
        output_dir = self.temp_dir / "smoke_output"
        _write_runtime_cards_workbook(
            xlsx_path,
            self.builder._fixture_cards(),
            self.mapper.FIELD_MAP,
            self.xlsx_export.PROFILES["decklists"],
            self.xlsx_export.PROFILES["lookups_runtime"],
            include_decklists=False,
            include_lookups_runtime=False,
        )

        summary = self.runner.run_smoke(xlsx_path=xlsx_path, output_dir=output_dir)

        self.assertEqual(summary["exported_card_rows"], 5)
        self.assertTrue(summary["export_runtime_jsonl_exists"])
        self.assertTrue(summary["cards_jsonl_exists"])
        self.assertTrue(summary["manifest_exists"])
        self.assertTrue(summary["diagnostics_exists"])
        self.assertFalse(summary["validation_blocking"])
        self.assertEqual(summary["deck_reference_errors"], 0)
        self.assertEqual(summary["cards_source"], "export-derived")
        self.assertEqual(summary["decks_source"], "sample fixture")
        self.assertIn("decks", summary["fixture_components"])
        self.assertTrue((output_dir / "exports" / "EXPORT_RUNTIME.jsonl").is_file())
        self.assertTrue((output_dir / "runtime_package" / "cards.jsonl").is_file())

        manifest = json.loads((output_dir / "runtime_package" / "manifest.json").read_text(encoding="utf-8"))
        self.assertEqual(manifest["source_files"][1]["type"], "export_runtime_cards_jsonl")

    def test_smoke_runner_can_include_real_decklists(self):
        xlsx_path = self.temp_dir / "aeterna_cards_and_decks.xlsx"
        output_dir = self.temp_dir / "smoke_output_with_decks"
        _write_runtime_cards_workbook(
            xlsx_path,
            self.builder._fixture_cards(),
            self.mapper.FIELD_MAP,
            self.xlsx_export.PROFILES["decklists"],
            self.xlsx_export.PROFILES["lookups_runtime"],
            include_decklists=True,
            include_lookups_runtime=True,
        )

        summary = self.runner.run_smoke(xlsx_path=xlsx_path, output_dir=output_dir, include_decklists=True)

        self.assertEqual(summary["exported_card_rows"], 5)
        self.assertEqual(summary["decklist_export_rows"], 5)
        self.assertTrue(summary["decklist_export_jsonl_exists"])
        self.assertTrue(summary["validation_blocking"])
        self.assertEqual(summary["deck_reference_errors"], 0)
        self.assertEqual(summary["cards_source"], "export-derived")
        self.assertEqual(summary["decks_source"], "export-derived")
        self.assertNotIn("decks", summary["fixture_components"])
        self.assertTrue((output_dir / "exports" / "PRODUCT_DECKLISTS.jsonl").is_file())

        manifest = json.loads((output_dir / "runtime_package" / "manifest.json").read_text(encoding="utf-8"))
        self.assertEqual(manifest["source_files"][1]["type"], "export_runtime_cards_jsonl")
        self.assertEqual(manifest["source_files"][2]["type"], "product_decklists_jsonl")

    def test_smoke_runner_can_include_real_runtime_lookups(self):
        xlsx_path = self.temp_dir / "aeterna_cards_decks_lookups.xlsx"
        lookups_xlsx_path = self.temp_dir / "LOOKUPS.xlsx"
        output_dir = self.temp_dir / "smoke_output_with_lookups"
        _write_runtime_cards_workbook(
            xlsx_path,
            self.builder._fixture_cards(),
            self.mapper.FIELD_MAP,
            self.xlsx_export.PROFILES["decklists"],
            self.xlsx_export.PROFILES["lookups_runtime"],
            include_decklists=True,
            include_lookups_runtime=True,
        )
        _write_lookups_workbook(lookups_xlsx_path)

        summary = self.runner.run_smoke(
            xlsx_path=xlsx_path,
            lookups_xlsx_path=lookups_xlsx_path,
            output_dir=output_dir,
            include_decklists=True,
            include_lookups_runtime=True,
        )

        self.assertEqual(summary["exported_card_rows"], 5)
        self.assertEqual(summary["decklist_export_rows"], 5)
        self.assertGreater(summary["lookups_export_rows"], 7)
        self.assertTrue(summary["lookups_export_jsonl_exists"])
        self.assertFalse(summary["validation_blocking"])
        self.assertEqual(summary["deck_reference_errors"], 0)
        self.assertEqual(summary["unknown_realm_errors"], 0)
        self.assertEqual(summary["unknown_card_type_errors"], 0)
        self.assertEqual(summary["diagnostic_count"], 0)
        self.assertEqual(summary["lookups_xlsx_path"], str(lookups_xlsx_path))
        self.assertEqual(summary["lookups_source"], "LOOKUPS.xlsx:RUNTIME_CORE+RUNTIME_ABILITY")
        self.assertEqual(summary["normalization_aliases_source"], "LOOKUPS.xlsx:RUNTIME_LEGACY_ALIAS")
        self.assertEqual(summary["normalization_aliases_count"], 2)
        self.assertEqual(summary["normalization_aliases_requires_audit_count"], 1)
        self.assertEqual(summary["normalization_aliases_allowed_count"], 1)
        self.assertEqual(summary["normalization_audit_matches"], 0)
        self.assertEqual(summary["normalization_audit_requires_audit"], 0)
        self.assertEqual(summary["normalization_audit_allowed"], 0)
        self.assertEqual(summary["normalization_preview_items"], 0)
        self.assertEqual(summary["normalization_preview_skipped_requires_audit"], 0)
        self.assertEqual(summary["normalization_preview_applied"], 0)
        self.assertEqual(summary["normalization_patch_plan_ready"], 0)
        self.assertEqual(summary["normalization_patch_plan_blocked"], 0)
        self.assertEqual(summary["normalization_patch_plan_applied"], 0)
        self.assertFalse(summary["normalization_apply_enabled"])
        self.assertEqual(summary["normalization_apply_applied"], 0)
        self.assertEqual(summary["normalization_apply_skipped"], 0)
        self.assertEqual(summary["normalization_apply_conflicts"], 0)
        self.assertNotIn("lookups", summary["fixture_components"])

        manifest = json.loads((output_dir / "runtime_package" / "manifest.json").read_text(encoding="utf-8"))
        self.assertEqual(manifest["source_files"][1]["type"], "lookups_runtime_jsonl")
        self.assertEqual(manifest["source_files"][2]["type"], "export_runtime_cards_jsonl")
        self.assertEqual(manifest["source_files"][3]["type"], "product_decklists_jsonl")
        self.assertEqual(manifest["source_files"][4]["type"], "lookups_xlsx_runtime_legacy_aliases")
        manifest_files = {item["path"] for item in manifest["files"]}
        self.assertIn("normalization_aliases.json", manifest_files)
        self.assertIn("normalization_audit_report.json", manifest_files)
        self.assertIn("normalization_preview_report.json", manifest_files)
        self.assertIn("normalization_patch_plan.json", manifest_files)
        self.assertIn("normalization_apply_report.json", manifest_files)
        normalization_aliases = json.loads(
            (output_dir / "runtime_package" / "normalization_aliases.json").read_text(encoding="utf-8")
        )
        self.assertEqual(len(normalization_aliases["normalization_aliases"]), 2)
        self.assertTrue(normalization_aliases["normalization_aliases"][1]["requires_audit"])
        self.assertFalse(normalization_aliases["normalization_aliases"][1]["normalization_allowed"])
        normalization_audit = json.loads(
            (output_dir / "runtime_package" / "normalization_audit_report.json").read_text(encoding="utf-8")
        )
        self.assertEqual(normalization_audit["summary"]["matches_total"], 0)
        normalization_preview = json.loads(
            (output_dir / "runtime_package" / "normalization_preview_report.json").read_text(encoding="utf-8")
        )
        self.assertEqual(normalization_preview["summary"]["preview_items"], 0)
        self.assertEqual(normalization_preview["summary"]["applied"], 0)
        patch_plan = json.loads(
            (output_dir / "runtime_package" / "normalization_patch_plan.json").read_text(encoding="utf-8")
        )
        self.assertEqual(patch_plan["summary"]["patches_ready"], 0)
        self.assertEqual(patch_plan["summary"]["applied"], 0)
        apply_report = json.loads(
            (output_dir / "runtime_package" / "normalization_apply_report.json").read_text(encoding="utf-8")
        )
        self.assertFalse(apply_report["summary"]["enabled"])
        self.assertEqual(apply_report["summary"]["applied"], 0)


def _write_runtime_cards_workbook(
    path,
    sample_cards,
    field_map,
    decklists_profile,
    lookups_profile,
    include_decklists=False,
    include_lookups_runtime=False,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "7. EXPORT_RUNTIME"

    headers = list(field_map.keys())
    for extra_header in ("Set_ID", "Collector_Number"):
        if extra_header not in headers:
            headers.append(extra_header)
    worksheet.append(headers)

    for index, card in enumerate(sample_cards, start=1):
        values_by_target = {
            "card_id": card["card_id"],
            "name_hu": card["name_hu"],
            "card_type": _real_card_type(card["card_type"]) if include_lookups_runtime else card["card_type"],
            "realm": _real_realm(card["realm"]) if include_lookups_runtime else card["realm"],
            "clan": card["clan"],
            "species": "none",
            "class": "none",
            "magnitude": card["magnitude"],
            "aura_cost": card["aura_cost"],
            "atk": card["atk"],
            "hp": card["hp"],
            "text_hu": "Sample XLSX ability text.",
            "structured_ability": "sample_module()",
            "recognized_zone": "none",
            "keywords": ";".join(card["keywords"]),
            "trigger": "none",
            "target": "none",
            "effect_tags": "sample",
            "duration": "none",
            "condition": "none",
            "machine_description": "Sample XLSX machine text.",
            "interpretation_status": card["interpretation_status"],
            "engine_notes": "Real export runtime package smoke.",
        }
        row = []
        for header in headers:
            if header == "Set_ID":
                row.append("SAMPLE")
            elif header == "Collector_Number":
                row.append(index)
            else:
                row.append(values_by_target[field_map[header]])
        worksheet.append(row)

    if include_decklists:
        deck_sheet = workbook.create_sheet("15. PRODUCT_DECKLISTS")
        headers = list(decklists_profile.required_fields)
        for optional_header in ("Kártya_Név", "Szerep_A_Pakliban", "Megjegyzés"):
            if optional_header not in headers:
                headers.append(optional_header)
        deck_sheet.append(headers)
        for card in sample_cards:
            values = {
                "Product_ID": "TEST-CORE01-IGNIS",
                "Deck_ID": "DECK-IGN-HAM-TEST-001",
                "Card_ID": card["card_id"],
                "Kártya_Név": card["name_hu"],
                decklists_profile.required_fields[3]: 1,
                "Szerep_A_Pakliban": "smoke",
                "Megjegyzés": "temporary smoke fixture",
            }
            deck_sheet.append([values.get(header, "none") for header in headers])

    if include_lookups_runtime:
        lookup_sheet = workbook.create_sheet("5A. LOOKUPS_RUNTIME")
        lookup_headers = list(lookups_profile.output_fields)
        lookup_sheet.append(lookup_headers)
        for row in _runtime_lookup_rows():
            lookup_sheet.append([row.get(header, "none") for header in lookup_headers])

    workbook.save(path)
    workbook.close()


def _real_card_type(value):
    return {
        "Entitas": "Entitás",
        "Rituale": "Rituálé",
        "Sik": "Sík",
    }.get(value, value)


def _real_realm(value):
    return {
        "Ignis": "IGNIS",
        "Aqua": "AQUA",
    }.get(value, value)


def _runtime_lookup_rows():
    values = [
        ("Card_Type", "Entitás"),
        ("Card_Type", "Ige"),
        ("Card_Type", "Rituálé"),
        ("Card_Type", "Jel"),
        ("Card_Type", "Sík"),
        ("Realm", "IGNIS"),
        ("Realm", "AQUA"),
    ]
    return [
        {
            "Lookup_Group": group,
            "Value": value,
            "Label_HU": value,
            "Status": "active",
            "Canonical_Value": value,
            "Used_For": "runtime_validation",
            "Sort_Order": index,
            "Source": "test",
            "Notes": "temporary smoke fixture",
        }
        for index, (group, value) in enumerate(values, start=1)
    ]


def _write_lookups_workbook(path):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    core_sheet = workbook.create_sheet("RUNTIME_CORE")
    ability_sheet = workbook.create_sheet("RUNTIME_ABILITY")
    legacy_alias_sheet = workbook.create_sheet("RUNTIME_LEGACY_ALIAS")
    headers = [
        "Lookup_Group",
        "Value",
        "Label_HU",
        "Status",
        "Canonical_Value",
        "Used_For",
        "Sort_Order",
        "Source",
        "Notes",
    ]
    core_sheet.append(headers)
    ability_sheet.append(headers)
    legacy_alias_sheet.append(headers)
    for row in _runtime_lookup_rows():
        target_sheet = ability_sheet if row["Lookup_Group"] in ("Keyword", "Trigger", "Target", "Effect_Tag", "Duration") else core_sheet
        row = dict(row)
        row["Value"] = _canonical_lookup_value(row["Lookup_Group"], row["Value"])
        row["Canonical_Value"] = row["Value"]
        target_sheet.append([row.get(header, "none") for header in headers])
    for row in _runtime_legacy_alias_rows():
        legacy_alias_sheet.append([row.get(header, "none") for header in headers])
    workbook.save(path)
    workbook.close()


def _canonical_lookup_value(group, value):
    if group == "Card_Type":
        return {
            "EntitĂˇs": "entity",
            "Ige": "incantation",
            "RituĂˇlĂ©": "ritual",
            "Jel": "sigil",
            "SĂ­k": "plane",
        }.get(value, value)
    if group == "Realm":
        return str(value).lower()
    return value


def _runtime_legacy_alias_rows():
    return [
        {
            "Lookup_Group": "Card_Type",
            "Value": "tactic",
            "Label_HU": "taktika",
            "Status": "legacy",
            "Canonical_Value": "ritual",
            "Used_For": "runtime_normalization",
            "Sort_Order": 10,
            "Source": "test",
            "Notes": "temporary smoke fixture",
        },
        {
            "Lookup_Group": "Card_Type",
            "Value": "spell",
            "Label_HU": "varazslat",
            "Status": "legacy",
            "Canonical_Value": "audit_required",
            "Used_For": "runtime_normalization",
            "Sort_Order": 20,
            "Source": "test",
            "Notes": "temporary smoke fixture",
        },
    ]


if __name__ == "__main__":
    unittest.main()
