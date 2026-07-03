import importlib.util
import io
import sys
import unittest
from pathlib import Path

from openpyxl import Workbook


SCRIPT_PATH = (
    Path(__file__).resolve().parents[1]
    / "tools"
    / "runtime_package"
    / "lookups_xlsx_reader.py"
)


def _load_reader_module():
    spec = importlib.util.spec_from_file_location("lookups_xlsx_reader", SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    sys.modules["lookups_xlsx_reader"] = module
    spec.loader.exec_module(module)
    return module


class TestLookupsXlsxReader(unittest.TestCase):
    def setUp(self):
        self.reader = _load_reader_module()

    def test_reads_runtime_core_and_ability_sheets(self):
        workbook_bytes = _build_workbook(
            runtime_core_rows=[
                _lookup_row("Card_Type", "entity", "runtime_validation", 10.0),
                _lookup_row("Realm", "", "runtime_validation", 20),
                _lookup_row("Realm", "old_realm", "runtime_validation", 30, status="deprecated"),
            ],
            runtime_ability_rows=[
                _lookup_row("Effect_Tag", "damage", "runtime_validation; ability_validation", 10),
                _lookup_row("Target", "enemy, entity", "runtime_validation", 20),
            ],
        )

        result = self.reader.load_runtime_lookups_from_xlsx(workbook_bytes)

        self.assertEqual(result["summary"]["sheets_read"], ["RUNTIME_CORE", "RUNTIME_ABILITY"])
        self.assertEqual(result["summary"]["records_read"], 5)
        self.assertEqual(result["summary"]["lookups_loaded"], 3)
        self.assertEqual(result["summary"]["skipped_empty_value"], 1)
        self.assertEqual(result["summary"]["skipped_inactive"], 1)

        lookups = result["lookups"]
        self.assertEqual(lookups[0]["lookup_group"], "card_type")
        self.assertEqual(lookups[0]["value"], "entity")
        self.assertEqual(lookups[0]["sort_order"], 10)
        self.assertEqual(lookups[1]["lookup_group"], "effect_tag")
        self.assertEqual(lookups[1]["used_for"], ["runtime_validation", "ability_validation"])
        self.assertEqual(lookups[2]["lookup_group"], "target")
        self.assertEqual(lookups[2]["used_for"], ["runtime_validation"])

    def test_missing_runtime_sheet_is_error(self):
        workbook = Workbook()
        workbook.active.title = "RUNTIME_CORE"
        workbook.active.append(self.reader.LOOKUP_HEADERS)
        workbook_bytes = io.BytesIO()
        workbook.save(workbook_bytes)
        workbook_bytes.seek(0)
        workbook.close()

        with self.assertRaises(self.reader.LookupsXlsxReaderError):
            self.reader.load_runtime_lookups_from_xlsx(workbook_bytes)


def _build_workbook(runtime_core_rows, runtime_ability_rows):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for sheet_name, rows in (
        ("RUNTIME_CORE", runtime_core_rows),
        ("RUNTIME_ABILITY", runtime_ability_rows),
    ):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(
            [
                "Lookup_Group",
                "Value",
                "Label_HU",
                "Status",
                "Canonical_Value",
                "Used_For",
                "Sort_Order",
                "Source",
                "Notes",
            ]
        )
        for row in rows:
            sheet.append(row)

    workbook_bytes = io.BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    workbook.close()
    return workbook_bytes


def _lookup_row(group, value, used_for, sort_order, status="active"):
    return [
        group,
        value,
        value,
        status,
        value,
        used_for,
        sort_order,
        "test",
        "in-memory fixture",
    ]


if __name__ == "__main__":
    unittest.main()
