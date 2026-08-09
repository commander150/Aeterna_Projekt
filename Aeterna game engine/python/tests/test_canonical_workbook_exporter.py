import importlib.util
import json
import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "tools" / "canonical_export" / "canonical_workbook_exporter.py"


def load_exporter_module():
    spec = importlib.util.spec_from_file_location("canonical_workbook_exporter", SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


exporter = load_exporter_module()


class TestCanonicalWorkbookExporter(unittest.TestCase):
    def setUp(self):
        self._temporary_directory = tempfile.TemporaryDirectory(prefix="aeterna_canonical_export_test_")
        self.root = Path(self._temporary_directory.name)
        self.workbook_path = self.root / "TEST.xlsx"
        self._write_workbook(self.workbook_path)

    def tearDown(self):
        self._temporary_directory.cleanup()

    def test_manifest_order_enabled_tables_and_exact_filenames(self):
        package = exporter.export_canonical_workbooks([self.workbook_path], self.root / "out")[0]

        self.assertEqual(
            {path.name for path in package.iterdir()},
            {"test.export_manifest.json", "test.meta.json", "test.schema_tables.json", "test.schema_fields.json", "test.validation_rules.json", "declared.items.json"},
        )
        self.assertFalse((package / "must-not-exist.json").exists())
        manifest = json.loads((package / "test.export_manifest.json").read_text(encoding="utf-8"))
        self.assertEqual(manifest["table_id"], "export_manifest")
        self.assertEqual(
            [record["table_id"] for record in manifest["records"] if record["export_enabled"]],
            ["export_manifest", "meta", "schema_tables", "schema_fields", "validation_rules", "items"],
        )
        self.assertEqual(manifest["records"][5]["export_file"], "declared.items.json")

    def test_export_is_byte_deterministic(self):
        first = exporter.export_canonical_workbooks([self.workbook_path], self.root / "one")[0]
        second = exporter.export_canonical_workbooks([self.workbook_path], self.root / "two")[0]

        first_bytes = {path.name: path.read_bytes() for path in first.iterdir()}
        second_bytes = {path.name: path.read_bytes() for path in second.iterdir()}
        self.assertEqual(first_bytes, second_bytes)

    def test_nullable_null_sentinel_becomes_json_null_and_all_fields_remain(self):
        package = exporter.export_canonical_workbooks([self.workbook_path], self.root / "out")[0]
        items = json.loads((package / "declared.items.json").read_text(encoding="utf-8"))
        meta = json.loads((package / "test.meta.json").read_text(encoding="utf-8"))

        self.assertIn("note", items["records"][0])
        self.assertIsNone(items["records"][0]["note"])
        values = {record["key"]: record["value"] for record in meta["records"]}
        self.assertEqual(values["null_sentinel"], "#NULL")
        self.assertEqual(values["tbd_sentinel"], "#TBD")

    def test_forbidden_null_is_blocking(self):
        self._set_cell("ITEMS", "C2", "#NULL")
        self._assert_code("CANONICAL_NULL_FORBIDDEN")

    def test_production_tbd_is_blocking(self):
        self._set_cell("ITEMS", "C2", "#TBD")
        self._assert_code("CANONICAL_TBD_FORBIDDEN")

    def test_development_tbd_is_preserved(self):
        self._set_cell("ITEMS", "C2", "#TBD")

        package = exporter.export_canonical_workbooks(
            [self.workbook_path], self.root / "out", production=False
        )[0]
        items = json.loads((package / "declared.items.json").read_text(encoding="utf-8"))

        self.assertEqual(items["records"][0]["value"], "#TBD")

    def test_missing_null_sentinel_is_blocking(self):
        self._delete_meta_key("null_sentinel")
        self._assert_code("CANONICAL_NULL_SENTINEL_INVALID")

    def test_missing_tbd_sentinel_is_blocking(self):
        self._delete_meta_key("tbd_sentinel")
        self._assert_code("CANONICAL_TBD_SENTINEL_INVALID")

    def test_invalid_null_sentinel_is_blocking(self):
        self._set_meta_value("null_sentinel", "NULL")
        self._assert_code("CANONICAL_NULL_SENTINEL_INVALID")

    def test_invalid_tbd_sentinel_is_blocking(self):
        self._set_meta_value("tbd_sentinel", "TBD")
        self._assert_code("CANONICAL_TBD_SENTINEL_INVALID")

    def test_export_filename_contract_is_exact_and_platform_independent(self):
        cases = {
            "valid.json": True,
            "invalid.JSON": False,
            "../escape.json": False,
            "sub/file.json": False,
            r"C:\escape.json": False,
            "/escape.json": False,
        }
        for index, (value, valid) in enumerate(cases.items()):
            with self.subTest(value=value):
                workbook_path = self.root / f"filename_{index}.xlsx"
                self._write_workbook(workbook_path)
                self._set_workbook_cell(workbook_path, "EXPORT_MANIFEST", "C7", value)
                if valid:
                    package = exporter.export_canonical_workbooks(
                        [workbook_path], self.root / f"filename_out_{index}"
                    )[0]
                    self.assertTrue((package / value).is_file())
                else:
                    self._assert_workbook_code(
                        workbook_path,
                        "CANONICAL_EXPORT_FILE_INVALID",
                        self.root / f"filename_out_{index}",
                    )

    def test_export_directory_name_must_be_a_direct_output_child(self):
        invalid_values = ["", "   ", ".", "..", "../escape", "sub/package", r"C:\escape", "/escape"]
        for index, value in enumerate(invalid_values):
            with self.subTest(value=value):
                workbook_path = self.root / f"directory_{index}.xlsx"
                self._write_workbook(workbook_path)
                self._set_workbook_cell(workbook_path, "META", "B5", value)
                self._assert_workbook_code(
                    workbook_path,
                    "CANONICAL_PACKAGE_DIRECTORY_INVALID",
                    self.root / f"directory_out_{index}",
                )

        for index, value in enumerate(("REGISTRY", "CARDDATABASE")):
            with self.subTest(value=value):
                workbook_path = self.root / f"valid_directory_{index}.xlsx"
                self._write_workbook(workbook_path, directory_name=value)
                package = exporter.export_canonical_workbooks(
                    [workbook_path], self.root / f"valid_directory_out_{index}"
                )[0]
                self.assertEqual(package, self.root / f"valid_directory_out_{index}" / value)

    def test_missing_manifest_referenced_sheet_is_blocking(self):
        self._set_cell("SCHEMA_TABLES", "B7", "MISSING")
        self._assert_code("CANONICAL_EXPORTED_SHEET_MISSING")

    def test_header_mismatch_is_blocking(self):
        self._set_cell("ITEMS", "B1", "wrong_header")
        self._assert_code("CANONICAL_HEADER_MISMATCH")

    def test_duplicate_field_id_is_blocking(self):
        workbook = load_workbook(self.workbook_path)
        try:
            sheet = workbook["SCHEMA_FIELDS"]
            sheet.cell(sheet.max_row, 1, sheet.cell(sheet.max_row - 1, 1).value)
            workbook.save(self.workbook_path)
        finally:
            workbook.close()
        self._assert_code("CANONICAL_FIELD_ID_DUPLICATE")

    def test_duplicate_table_field_is_blocking(self):
        workbook = load_workbook(self.workbook_path)
        try:
            sheet = workbook["SCHEMA_FIELDS"]
            sheet.cell(sheet.max_row, 2, sheet.cell(sheet.max_row - 1, 2).value)
            sheet.cell(sheet.max_row, 3, sheet.cell(sheet.max_row - 1, 3).value)
            workbook.save(self.workbook_path)
        finally:
            workbook.close()
        self._assert_code("CANONICAL_TABLE_FIELD_DUPLICATE")

    def test_duplicate_primary_record_id_is_blocking(self):
        workbook = load_workbook(self.workbook_path)
        try:
            sheet = workbook["ITEMS"]
            sheet.append(["item_1", "second", "value"])
            workbook.save(self.workbook_path)
        finally:
            workbook.close()
        self._assert_code("CANONICAL_PRIMARY_KEY_DUPLICATE")

    def test_blocked_export_does_not_replace_existing_package(self):
        existing = self.root / "out" / "TEST_PACKAGE"
        existing.mkdir(parents=True)
        marker = existing / "marker.txt"
        marker.write_text("previous", encoding="utf-8")
        self._set_cell("ITEMS", "C2", "#TBD")

        self._assert_code("CANONICAL_TBD_FORBIDDEN", output_root=self.root / "out")
        self.assertEqual(marker.read_text(encoding="utf-8"), "previous")
        self.assertEqual(list(existing.iterdir()), [marker])

    def test_multi_package_publish_rolls_back_as_one_transaction(self):
        first_workbook = self.root / "FIRST.xlsx"
        second_workbook = self.root / "SECOND.xlsx"
        self._write_workbook(first_workbook, package_id="aeterna_first", directory_name="FIRST_PACKAGE", file_prefix="first")
        self._write_workbook(second_workbook, package_id="aeterna_second", directory_name="SECOND_PACKAGE", file_prefix="second")
        output_root = self.root / "transaction"
        first_target = output_root / "FIRST_PACKAGE"
        second_target = output_root / "SECOND_PACKAGE"
        first_target.mkdir(parents=True)
        second_target.mkdir(parents=True)
        (first_target / "old.txt").write_text("first-old", encoding="utf-8")
        (second_target / "old.txt").write_text("second-old", encoding="utf-8")
        real_replace = os.replace

        def fail_second_publish(source, destination):
            if Path(source).name == "SECOND_PACKAGE" and Path(destination) == second_target:
                raise OSError("injected second-package publish failure")
            return real_replace(source, destination)

        with mock.patch.object(exporter.os, "replace", side_effect=fail_second_publish):
            with self.assertRaisesRegex(OSError, "injected second-package"):
                exporter.export_canonical_workbooks([first_workbook, second_workbook], output_root)

        self.assertEqual({path.name for path in first_target.iterdir()}, {"old.txt"})
        self.assertEqual({path.name for path in second_target.iterdir()}, {"old.txt"})
        self.assertEqual((first_target / "old.txt").read_text(encoding="utf-8"), "first-old")
        self.assertEqual((second_target / "old.txt").read_text(encoding="utf-8"), "second-old")
        self.assertFalse(any(path.name.startswith(".canonical-export-") for path in output_root.iterdir()))

    def test_multi_package_rollback_removes_originally_absent_package(self):
        first_workbook = self.root / "NEW.xlsx"
        second_workbook = self.root / "EXISTING.xlsx"
        self._write_workbook(first_workbook, package_id="aeterna_new", directory_name="NEW_PACKAGE", file_prefix="new")
        self._write_workbook(second_workbook, package_id="aeterna_existing", directory_name="EXISTING_PACKAGE", file_prefix="existing")
        output_root = self.root / "absent_transaction"
        existing_target = output_root / "EXISTING_PACKAGE"
        existing_target.mkdir(parents=True)
        (existing_target / "old.txt").write_text("existing-old", encoding="utf-8")
        new_target = output_root / "NEW_PACKAGE"
        real_replace = os.replace

        def fail_second_publish(source, destination):
            if Path(source).name == "EXISTING_PACKAGE" and Path(destination) == existing_target:
                raise OSError("injected existing-package publish failure")
            return real_replace(source, destination)

        with mock.patch.object(exporter.os, "replace", side_effect=fail_second_publish):
            with self.assertRaisesRegex(OSError, "injected existing-package"):
                exporter.export_canonical_workbooks([first_workbook, second_workbook], output_root)

        self.assertFalse(new_target.exists())
        self.assertEqual({path.name for path in existing_target.iterdir()}, {"old.txt"})
        self.assertEqual((existing_target / "old.txt").read_text(encoding="utf-8"), "existing-old")
        self.assertFalse(any(path.name.startswith(".canonical-export-") for path in output_root.iterdir()))

    def test_successful_publish_leaves_no_staging_or_backup(self):
        output_root = self.root / "clean_success"
        exporter.export_canonical_workbooks([self.workbook_path], output_root)

        self.assertFalse(any(path.name.startswith(".canonical-export-") for path in output_root.iterdir()))

    def test_export_does_not_require_repository_temp_directory(self):
        clean_checkout = self.root / "clean_checkout"
        clean_checkout.mkdir()
        previous_working_directory = Path.cwd()
        try:
            os.chdir(clean_checkout)
            package = exporter.export_canonical_workbooks(
                [self.workbook_path], self.root / "outside_checkout_output"
            )[0]
        finally:
            os.chdir(previous_working_directory)

        self.assertTrue(package.is_dir())
        self.assertFalse((clean_checkout / "TEMP").exists())

    def test_structured_diagnostics_name_workbook_sheet_row_field_and_rule(self):
        self._set_cell("ITEMS", "C2", "#TBD")
        report = self.root / "diagnostics.json"
        with self.assertRaises(exporter.CanonicalExportError):
            exporter.export_canonical_workbooks([self.workbook_path], self.root / "out", diagnostics_path=report)

        diagnostic = json.loads(report.read_text(encoding="utf-8"))["diagnostics"][0]
        self.assertEqual(diagnostic["workbook"], "TEST.xlsx")
        self.assertEqual(diagnostic["sheet"], "ITEMS")
        self.assertEqual(diagnostic["row"], 2)
        self.assertEqual(diagnostic["field"], "value")
        self.assertEqual(diagnostic["rule"], "production_tbd_policy")

    def _assert_code(self, code, output_root=None):
        self._assert_workbook_code(self.workbook_path, code, output_root or self.root / "out")

    def _assert_workbook_code(self, workbook_path, code, output_root):
        with self.assertRaises(exporter.CanonicalExportError) as raised:
            exporter.export_canonical_workbooks([workbook_path], output_root)
        self.assertIn(code, [item.code for item in raised.exception.diagnostics])

    def _set_cell(self, sheet_name, coordinate, value):
        self._set_workbook_cell(self.workbook_path, sheet_name, coordinate, value)

    @staticmethod
    def _set_workbook_cell(workbook_path, sheet_name, coordinate, value):
        workbook = load_workbook(workbook_path)
        try:
            workbook[sheet_name][coordinate] = value
            workbook.save(workbook_path)
        finally:
            workbook.close()

    def _set_meta_value(self, key, value):
        workbook = load_workbook(self.workbook_path)
        try:
            sheet = workbook["META"]
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row, 1).value == key:
                    sheet.cell(row, 2).value = value
                    break
            else:
                raise AssertionError(f"META fixture key not found: {key}")
            workbook.save(self.workbook_path)
        finally:
            workbook.close()

    def _delete_meta_key(self, key):
        workbook = load_workbook(self.workbook_path)
        try:
            sheet = workbook["META"]
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row, 1).value == key:
                    sheet.delete_rows(row)
                    break
            else:
                raise AssertionError(f"META fixture key not found: {key}")
            workbook.save(self.workbook_path)
        finally:
            workbook.close()

    @staticmethod
    def _write_workbook(path, *, package_id="aeterna_test", directory_name="TEST_PACKAGE", file_prefix="test"):
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheets = {
            "META": ["key", "value", "value_type", "description"],
            "EXPORT_MANIFEST": ["table_id", "export_enabled", "export_file", "export_format", "export_order", "notes"],
            "SCHEMA_TABLES": ["table_id", "sheet_name", "record_type", "schema_version", "primary_key", "status", "notes"],
            "SCHEMA_FIELDS": ["field_id", "table_id", "field_name", "column_order", "data_type", "required_mode", "nullable", "null_handling", "tbd_allowed", "allowed_group_id", "reference_table_id", "reference_field_id", "default_value", "status", "notes"],
            "VALIDATION_RULES": ["validation_rule_id", "rule_scope_id", "target_table_id", "target_field_id", "validation_kind_id", "condition_expression", "operator_id", "comparison_value", "minimum_value", "maximum_value", "reference_table_id", "reference_field_id", "severity_id", "blocking", "error_code", "message_key", "message", "validation_stage_id", "status", "source_id", "source_ref", "notes"],
            "ITEMS": ["item_id", "note", "value"],
            "DISABLED": ["disabled_id"],
        }
        for name, headers in sheets.items():
            worksheet = workbook.create_sheet(name)
            worksheet.append(headers)

        meta = workbook["META"]
        for row in [
            ["workbook_id", package_id, "string", "package identity"],
            ["schema_version", "1.0.0", "version", "schema"],
            ["data_version", "1.0.0", "version", "data"],
            ["export_directory_name", directory_name, "string", "directory"],
            ["export_manifest_file", f"{file_prefix}.export_manifest.json", "string", "manifest"],
            ["null_sentinel", "#NULL", "string", "literal null marker"],
            ["tbd_sentinel", "#TBD", "string", "literal tbd marker"],
        ]:
            meta.append(row)

        manifest_rows = [
            ["export_manifest", True, f"{file_prefix}.export_manifest.json", "json", 1, "manifest"],
            ["meta", True, f"{file_prefix}.meta.json", "json", 2, "meta"],
            ["schema_tables", True, f"{file_prefix}.schema_tables.json", "json", 3, "tables"],
            ["schema_fields", True, f"{file_prefix}.schema_fields.json", "json", 4, "fields"],
            ["validation_rules", True, f"{file_prefix}.validation_rules.json", "json", 5, "rules"],
            ["items", True, "declared.items.json", "json", 10, "items"],
            ["disabled", False, "#NULL", "#NULL", "#NULL", "disabled"],
        ]
        for row in manifest_rows:
            workbook["EXPORT_MANIFEST"].append(row)

        definitions = [
            ("meta", "META", "key"),
            ("export_manifest", "EXPORT_MANIFEST", "table_id"),
            ("schema_tables", "SCHEMA_TABLES", "table_id"),
            ("schema_fields", "SCHEMA_FIELDS", "field_id"),
            ("validation_rules", "VALIDATION_RULES", "validation_rule_id"),
            ("items", "ITEMS", "item_id"),
            ("disabled", "DISABLED", "disabled_id"),
        ]
        for table_id, sheet_name, primary_key in definitions:
            workbook["SCHEMA_TABLES"].append([table_id, sheet_name, "fixture", "1.0.0", primary_key, "active", "fixture"])

        nullable_fields = {
            ("export_manifest", "export_file"),
            ("export_manifest", "export_format"),
            ("export_manifest", "export_order"),
            ("items", "note"),
        }
        for field_name in ("allowed_group_id", "reference_table_id", "reference_field_id", "default_value"):
            nullable_fields.add(("schema_fields", field_name))
        for table_id, sheet_name, _ in definitions:
            for order, field_name in enumerate(sheets[sheet_name], 1):
                nullable = (table_id, field_name) in nullable_fields
                workbook["SCHEMA_FIELDS"].append([
                    f"fld_{table_id}_{field_name}", table_id, field_name, order, "string", "optional" if nullable else "always", nullable,
                    "explicit_null" if nullable else "forbidden", False, "#NULL", "#NULL", "#NULL", "#NULL", "active", "fixture",
                ])

        workbook["ITEMS"].append(["item_1", "#NULL", "value"])
        workbook.save(path)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
